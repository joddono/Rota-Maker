#module imports for the program
from tkinter import *
from tkinter import ttk
from customtkinter import *
from PIL import ImageTk, Image, ImageDraw, ImageGrab
from datetime import *
from email.message import EmailMessage
from dateutil.relativedelta import relativedelta, MO
from tkinter import Tk, StringVar, Label, scrolledtext, Button, Canvas, filedialog
from openpyxl import Workbook
from openpyxl.styles import Font
from cryptography.fernet import Fernet
import datetime
import os
import bcrypt
import sqlite3
import matplotlib
import tkinter as tk
import openpyxl
import random
import ssl
import json
import smtplib
import time
import hashlib
import re
import csv
import calendar
#counters to only allow them to open one new unique window at a time 
global counter1
counter1 = 0
global counter2
counter2 = 0
global view_counter
view_counter = 0
global create_counter
create_counter = 0
global publish_counter
publish_counter = 0
global info_counter
info_counter = 0
global account_settings_counter
account_settings_counter = 0
global next_week
next_week = 0

#conect to the database
conn = sqlite3.connect('Rota_Maker.db')
cursor = conn.cursor()
#command to create the "Stores" table
create_table2_command = """CREATE TABLE IF NOT EXISTS "Stores" (
	"StoreID"	INTEGER,
	"Store_Name"	TEXT,
	PRIMARY KEY("StoreID" AUTOINCREMENT)
)"""
#run the command
cursor.execute(create_table2_command)
#checks if there is already data in the "Stores" table
check_values_command = """SELECT COUNT(*) FROM Stores;"""
cursor.execute(check_values_command)
count = cursor.fetchone()[0]
#if "Stores" is empty
if int(count) == 0:
    #insert the store names into the "Stores" table
    insert_store_names = """INSERT INTO Stores (Store_Name) VALUES 
    ('South Kensington'),
    ('Battersea'),
    ('Hampstead'),
    ('East Dulwich'),
    ('Chiswick'),
    ('Stoke Newington'),
    ('Wimbledon');"""
    #run the command
    cursor.execute(insert_store_names)
    conn.commit()
#command to create the "Account" table
create_table1_command = """CREATE TABLE IF NOT EXISTS "Account" (
	"AccountID"	INTEGER,
	"Name"	TEXT,
	"Email"	TEXT,
	"Password"	TEXT,
	"Birthday"	TEXT,
	"Monday_Working"	BIT,
	"Tuesday_Working"	BIT,
	"Wednesday_Working"	BIT,
	"Thursday_Working"	BIT,
	"Friday_Working"	BIT,
	"Saturday_Working"	BIT,
	"Sunday_Working"	BIT,
	"Contract_Hours"	INTEGER,
	"Primary_Store"	INTEGER,
	"Closest_Store"	INTEGER,
	"Admin"	BIT,
	PRIMARY KEY("AccountID" AUTOINCREMENT),
	FOREIGN KEY("Primary_Store") REFERENCES "Stores"("StoreID"),
	FOREIGN KEY("Closest_Store") REFERENCES "Stores"("StoreID")
);"""
#run the command
cursor.execute(create_table1_command)
conn.commit()
#create the AccountID in the "Account" table if it doesn't exist
create_index_command = """CREATE INDEX IF NOT EXISTS "ascending_accounts" ON "Account" (
    "AccountID" ASC
); """
cursor.execute(create_index_command)
conn.commit()
#close the connection to the database
conn.close()

#first window
class first_window:
    def __init__(self, root):
        #initialize tkinter pop up
        self.root1 = root
        self.root1.title("Home Page")
        self.root1.geometry("700x600+610+150")
        self.root1.configure(bg="#272727")
        set_appearance_mode("Dark")
        self.root1.resizable(False, False)
        #Label of name of project
        title_label = CTkLabel(master=self.root1, text=" The Rota Maker", font=("Comic Sans", 47))
        title_label.place(relx = 0.5, rely=0.2, anchor="center")
        #Credits to maker
        self.created_by_label = CTkLabel(master=root, text="by James Oddono", font=("Arial", 16, "bold"))
        self.created_by_label.place(relx=0.41, rely=0.24)
        #Button that when clicked will open the login page
        login_button = CTkButton(master=self.root1, width=193, height=80, text="Log In", font=("Helvetica", 20) ,corner_radius=1000,
                                 fg_color="transparent", border_color="#0078d7", border_width=4, hover_color="#0078d7",command=self.login_button_command)
        login_button.place(relx=0.3,rely=0.425, anchor="center")
        #Button that when clicked will open the create account
        create_account_button = CTkButton(master=self.root1,width=193, height=80, text="Create Account", font=("Helvetica", 20) , corner_radius=1000, fg_color="transparent", border_color="#0078d7",
                                          border_width=4, hover_color="#0078d7",command=self.create_account_button_command)
        create_account_button.place(relx=0.7,rely=0.425, anchor="center")
        #Button that when clicked will quit the program
        close_button = CTkButton(master=self.root1, width=195, height=70, text="Close", font=("Helvetica", 24) , corner_radius=1000, fg_color="#000000", hover_color="#F90101", command=self.root1.destroy)
        close_button.place(relx=0.5,rely=0.7, anchor="center")

    def login_button_command(self):
        #Restrict the amount of windows they can open    
        global counter1
        if counter1 < 1:
            #hide the main window
            self.root1.withdraw()
            #open new window by calling the login_window class
            self.root2 = CTkToplevel(self.root1)
            login_window(self.root2)
            #there is 1 opened window so counter is 1
            counter1 += 1

    def create_account_button_command(self):
        #Restrict the amount of windows they can open    
        global counter1
        if counter1 < 1:
            #hide the main window
            self.root1.withdraw()
            #open new window by calling the login_window class
            self.root3 = CTkToplevel(self.root1)
            create_account_window(self.root3)
            #there is 1 opened window so counter is 1
            counter1 += 1

#log in window
class login_window:
    def __init__(self, root):
        #initialize the tkinter pop up
        self.root2 = root
        self.root2.title("Log In")
        self.root2.geometry("420x500+755+200")
        self.root2.configure(bg="#272727")
        self.root2.resizable(False, False)
        #login label
        login_label = CTkLabel(master=root, text="Log in to your account", font=("Helvetica", 27))
        login_label.place(relx=0.5, rely=0.15, anchor="center")
        #email and password entry
        self.email_entry = CTkEntry(master=root, font=("Helvetica", 15), height = 35, placeholder_text="Email", width = 300)
        self.password_entry = CTkEntry(master=root, font=("Helvetica", 15), show="*" , height = 35, placeholder_text="Password", width = 300) 
        #show/hide icon for passwords
        self.show_password_img = CTkImage(Image.open("show_password_img.png"), size=(20, 20))
        self.hide_password_img = CTkImage(Image.open("hide_password_img.png"), size=(20, 20))
        self.toggle_password_button = CTkButton(master=root, image=self.show_password_img, text="", width = 20, height = 30, command=self.toggle_password)
        self.toggle_password_button.place(x=370,y=210)
        #placing the entry boxes
        self.email_entry.place(relx = 0.5, rely = 0.35, anchor="center")
        self.password_entry.place(relx=0.5, rely=0.45, anchor="center")
        #forgot password button
        forgot_password_btn = CTkButton(master=root, text="Forgot Password?", font=("Helvetica", 15),  height=25, width=100, corner_radius=6,
                                        fg_color="transparent", hover_color="#DD1D00", command=lambda: [self.forgot_password()])
        forgot_password_btn.place(x=60, y= 255)
        #log in button, will run the "check login info" method
        loginbtn = CTkButton(master=root, text="Log In", width = 300, height = 45, corner_radius=6 , fg_color="#0078d7", font=("Helvetica", 18), command=self.check_login_info)
        loginbtn.place(relx=0.5, rely= 0.7, anchor="center")
        #back button, opens the home page
        back_button = CTkButton(master=root, text="Back", font=("Helvetica", 18), width = 300, height = 45, corner_radius=6, fg_color="#0078d7",
                                hover_color="#DD1D00", command=lambda: [root.destroy(), self.back_button()])
        back_button.place(relx=0.5, rely=0.81, anchor="center")

    def toggle_password(self):
        #checks if the entry is showing
        if self.password_entry.cget("show") == "":
            #chanes the image to the show icon
            self.toggle_password_button.configure(image=self.show_password_img)
            #hides the characters in the entry box
            self.password_entry.configure(show="*")
        else:
            #changes the image to the hid icon
            self.toggle_password_button.configure(image=self.hide_password_img)
            #shows the characters in the entry box
            self.password_entry.configure(show="")

    def forgot_password(self):
        #checks if there is already a window open
        global counter2
        if counter2 < 1:
            #calls a new class "reset password window"
            self.root4 = CTkToplevel(root)
            reset_password_window(self.root4, self.root2)
            #1 window open, so counter should be 1
            counter2 += 1
            #hides the log in window
            self.root2.iconify()

    def back_button(self):
        #resets amount of windows open to 0
        global counter1
        counter1 = counter1 - 1
        #shows the home page again
        root.deiconify()

#check if they entered anything
    def check_login_info(self):
        #fetches the data in the entry boxes
        self.login_email = self.email_entry.get()
        self.login_password = self.password_entry.get()
        #checks if there is any data in the entry boxes
        if self.login_email and self.login_password:
            #calls the "save login info" method
            self.save_login_info()
        else:
            #error message pop up
            empty_error = tk.messagebox.showerror(title="Error", message="Fill in the blank fields")
            
    def save_login_info(self):
        #puts the class attributes as temorary variables
        login_email = self.login_email
        login_password = self.login_password
        #connects to the database
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        #gets the email and password for where the email is the email they have entered
        cursor.execute("SELECT Email, Password FROM Account WHERE Email=(?)", (login_email,))
        accountdata = cursor.fetchall()
        #close connection to the database
        conn.close()
        #checks if there was any data in the database with that email
        if accountdata:
            #gets the email and password that was in the database
            account_email = accountdata[0][0]
            account_password = accountdata[0][1]
            #checks if the password in the database matches the hash of the password they have entered
            if bcrypt.checkpw(login_password.encode("utf-8"), account_password):
                #close the login window
                self.root2.destroy()
                #call the "login successful" method
                self.login_successful()
            else:
                #error message pop up
                tk.messagebox.showerror(title="Error", message="Incorrect Password")
        else:
            #error message pop up
            tk.messagebox.showerror(title="Error", message="Account does not exist")

    def login_successful(self):
        #hide the main window
        root.withdraw()
        #connect to database
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        #get the users admin level
        data = (cursor.execute("SELECT admin FROM Account WHERE Email=(?)", (self.login_email,))).fetchall()
        admin = int(data[0][0])
        conn.close()
        root5 = CTk()
        #if they are standard employee
        if admin in (0, 1):
            #open employee window
            employee_window(root5, self.login_email)
        #if they are a manager (assistant manager, store manager or rota manager)
        elif admin in (2, 3, 4):
            #open manager window
            manager_window(root5, self.login_email)
        else:
            #display error message
            tk.messagebox.showerror("Error", "There is an error with your account")
            #destroy main window
            root.destroy()
        root5.mainloop()

#reset password window
class reset_password_window:
    def __init__(self, root, root2):
        #setting the window settings
        self.root2 = root2
        self.root = root
        self.root.title("Reset Password")
        self.root.geometry("500x400+700+200")
        self.root.configure(bg="#272727")
        self.root.resizable(False, False)
        self.title = CTkLabel(master=self.root, text="Reset Password", font=("Helvetica", 25))
        self.title.place(x=165, y=40)
        #entry box for their email
        self.email_entry = CTkEntry(master=root,width= 350, placeholder_text="Email of your account", font=("Helvetica", 15))
        self.email_entry.place(x=70, y=150)
        #button to continue to next window, runs "continue clicked" method
        self.continue_button = CTkButton(master=root, text="Continue", font=("Helvetica", 15), height=30, width=100, command=lambda: [self.continue_clicked()], fg_color="#0078d7")
        self.continue_button.place(x=200, y=200)
        #button to close the window and go back to the previous window
        self.back_button = CTkButton(master=root, text="Back", font=("Helvetica", 15), width=200, height=40, command=lambda: [self.back_clicked(), root.destroy()], fg_color="#0078d7")
        self.back_button.place(x=150, y=320)

    def back_clicked(self):
        #shows the log in window
        global counter2
        counter2 = counter2 - 1
        self.root2.deiconify()

    def generate_reset_code(self):
        #generates a random six digit code
        self.reset_code = random.randint(100000, 1000000)

    def continue_clicked(self):
        #fetches the email they entered
        self.entered_email = self.email_entry.get()
        #checks if it is not empty
        if self.entered_email:
            #connects to database
            conn = sqlite3.connect('Rota_Maker.db')
            cursor = conn.cursor()
            #selects all the data in the database where the email matches the email they entered 
            query = f"SELECT * FROM Account WHERE Email = ?"
            cursor.execute(query, (self.entered_email,))
            result = cursor.fetchall()
            #closes the connection
            conn.close()
            #checks if the result is not empty
            if result:
                #calls the method to generate the code
                self.generate_reset_code()
                #calls the "move continue button" method
                self.move_continue_button()
                #calls the "send email" method
                self.send_email()
            else:
                #error message pop up
                tk.messagebox.showerror(title="Error", message="There is no account with this email")
        else:
            #error message pop up
            tk.messagebox.showerror(title="Error", message="Please enter an email")

    def move_continue_button(self):
        #changes window size and placement
        self.root.geometry("700x450+600+200")
        #displays confirmation to the user
        confirmation_label = CTkLabel(master=self.root, text=f"An email has been sent to [{self.entered_email}] with your reset code", font=("Helvetica", 15))
        confirmation_label.place(relx=0.5, rely= 0.25, anchor="center")
        #entry box for the code
        self.code_entry = CTkEntry(master=self.root, placeholder_text="Reset Code", font=("Helvetica", 15), width=250)
        self.code_entry.place(x=225, y=130)
        #move all the buttons on the window
        self.continue_button.place(x=300, y=260)
        self.continue_button.configure(text="Reset", command=lambda: [self.continue_clicked_again()])
        self.email_entry.place_forget()
        self.back_button.place(x=250, y=380)
        self.title.place(x=250, y=40)
        #new password entry box
        self.password_entry = CTkEntry(master=self.root, placeholder_text = "New Password", font=("Helvetica", 15), width = 400, show="*")
        self.password_entry.place(x=150, y= 170)
        #new confirm password entry box
        self.confirm_password_entry = CTkEntry(master=self.root, placeholder_text="Confirm New Password", font=("Helvetica", 15), width=400, show="*")
        self.confirm_password_entry.place(x=150, y=210)
        #icons for showing/hiding password
        self.show_password_img = CTkImage(Image.open("show_password_img.png"), size=(20, 20))
        self.hide_password_img = CTkImage(Image.open("hide_password_img.png"), size=(20, 20))
        self.toggle_password_button = CTkButton(master=self.root, image=self.show_password_img, text="", width = 20, height = 25, command=self.toggle_password)
        self.toggle_password_button.place(x=560,y=170)
        self.show_password_img2 = CTkImage(Image.open("show_password_img.png"), size=(20, 20))
        self.hide_password_img2 = CTkImage(Image.open("hide_password_img.png"), size=(20, 20))
        self.toggle_password_button2 = CTkButton(master=self.root, image=self.show_password_img2, text="", width = 20, height = 25, command=self.toggle_password2)
        self.toggle_password_button2.place(x=560, y=210)
        #button to resend the code to their email
        self.send_again_label = CTkLabel(master=self.root, text=("Didn't recieve it?"), font=("Helvetica", 15))
        self.send_again_label.place(x=187, y=319)
        self.send_again_button = CTkButton(master=self.root, text=("Try Again"), font=("Helvetica", 15), fg_color="transparent", hover_color="#DD1D00",
                                           height=25, width=100, command=lambda: [self.send_email()])
        self.send_again_button.place(x=300, y=320)

    def toggle_password(self):
        #checks if the characters are showing
        if self.password_entry.cget("show") == "":
            #changes the image icon to the 'show' version
            self.toggle_password_button.configure(image=self.show_password_img)
            #changes the entry box to hide the characters
            self.password_entry.configure(show="*")
        else:
            #changes the image icon to the 'hide' version
            self.toggle_password_button.configure(image=self.hide_password_img)
            #changes the entry box to show the characters
            self.password_entry.configure(show="")

    def toggle_password2(self):
        #checks if the characters are showing
        if self.confirm_password_entry.cget("show") == "":
            #changes the image icon to the 'show' version
            self.toggle_password_button2.configure(image=self.show_password_img2)
            #changes the entry box to hide the characters
            self.confirm_password_entry.configure(show="*")
        else:
            #changes the image icon to the 'hide' version
            self.toggle_password_button2.configure(image=self.hide_password_img2)
            #changes the entry box to show the characters
            self.confirm_password_entry.configure(show="")

    def continue_clicked_again(self):
        #gets the entered code from the entry box
        self.entered_reset_code = self.code_entry.get()
        #gets the entered password from the entry box
        self.entered_password = self.password_entry.get()
        #gets the entered confirm password from the entry box
        self.entered_confirm_password = self.confirm_password_entry.get()
        #checks if any of them are empty
        if self.entered_reset_code and self.entered_password and self.entered_confirm_password:
            #checks if the reset code was correct
            if int(self.entered_reset_code) == int(self.reset_code):
                #checks if the two password match
                if self.entered_password == self.entered_confirm_password:
                    #calls the 'update password' method
                    self.update_password()
                    #success pop up display
                    tk.messagebox.showinfo("Success", "Password reset successfully!")
                    #closes the forgot password window
                    self.root.destroy()
                    global counter2
                    counter2 = counter2 - 1
                    #shows the log in window
                    self.root2.deiconify()
                else:
                    tk.messagebox.showerror(title="Error", message="Passwords do not match")
            else:
                tk.messagebox.showerror(title="Error", message="Invalid reset code entered")
        else:
            tk.messagebox.showerror(title="Error", message="Please fill in the fields")

    def update_password(self):
        #hash the entered password
        self.hashed_password = bcrypt.hashpw((self.entered_password.encode("utf-8")), bcrypt.gensalt())
        #connect to database
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        #update the new hashed password with the old hashed password
        query = """UPDATE Account SET Password = ? WHERE Email = ?"""
        cursor.execute(query, (self.hashed_password, self.entered_email))
        conn.commit()
        #close connection
        conn.close()
        
    # Function to generate or retrieve the encryption key
    def get_key(self):
        # Check if the key file exists
        if os.path.exists('key.key'):
            # Read the key from the file
            with open('key.key', 'rb') as key_file:
                key = key_file.read()
        return key

    # Function to decrypt the password
    def decrypt_password(self, encrypted_password, key):
        # Create a Fernet cipher suite with the provided key
        cipher_suite = Fernet(key)
        # Decrypt the encrypted password and decode it
        decrypted_password = cipher_suite.decrypt(encrypted_password).decode()
        return decrypted_password

    # Function to retrieve encrypted password from JSON file
    def retrieve_encrypted_password(self):
        # Open the encrypted_password.json file for reading
        with open('encrypted_password.json', 'r') as file:
            # Load JSON data from the file
            data = json.load(file)
            # Extract the encrypted password from the loaded data
            encrypted_password = data['password']
        return encrypted_password

    def send_email(self):
        # Load configuration from the JSON file
        with open('encrypted_password.json', 'r') as config_file:
            config = json.load(config_file)
        #set email that will send the automated message (burner gmail)
        email_sender = ""
        #get the encrypted app password
        enc_pass = self.retrieve_encrypted_password()
        #get the encryption key (symmetric encryption)
        key = self.get_key()
        email_password = self.decrypt_password(enc_pass, key)
        #get the email they entered
        email_reciever = self.entered_email
        #set the email subject and message
        subject = "Password Reset Request"
        body = f"""
If you didn't make this request, ignore this email.

Your reset code is:
{self.reset_code}
        
Thank you for using The Rota Maker.

This email message was auto-generated.
Please do not respond.
        """
        #make a variable an object of the smtp module to use in sending the email remotely
        msg = EmailMessage()
        #set the 'from' as the sender email variable
        msg['From'] = email_sender
        #set the 'to' as the reciever email variable
        msg['To'] = email_reciever
        #set the subject as the subject variable
        msg['Subject'] = subject
        #set the email content as the body variable
        msg.set_content(body)            
        # Create a default SSL context to establish a secure connection to SMTP server
        context = ssl.create_default_context()
        try:
            # Connect to the SMTP server using SSL
            with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
                # Login to the SMTP server using sender email and password
                smtp.login(email_sender, email_password)
                # Send the email
                smtp.sendmail(email_sender, email_reciever, msg.as_string())
        except Exception as e:
            # If an error occurs during sending email, show an info dialog with the error message
            error_msg = (f"An error occurred: {e}")
            tk.messagebox.showinfo(title="Info", message=error_msg)

# Create account window class
class create_account_window:
    def __init__(self, root):
        # Initialize the window
        self.root = root
        self.root.title("Create Account")
        self.root.geometry("500x870+710+50")
        self.root.configure(bg="#272727")
        self.root.resizable(False, False)
        # Create entry fields for full name, email, and passwords
        self.fullname_entry = CTkEntry(master=root, font=("Helvetica", 18), height=35, width=380, placeholder_text="Forename and Surname")
        self.fullname_entry.place(x=60, y=45)
        self.email_entry = CTkEntry(master=root, font=("Helvetica", 18), height=35, width=380, placeholder_text="Email Address")
        self.email_entry.place(x=60, y=95)
        self.password_entry = CTkEntry(master=root, show="*", font=("Helvetica", 18), height=35, width=380, placeholder_text="Password")
        self.password_entry.place(x=60, y=145)
        self.confirm_password_entry = CTkEntry(master=root, show="*", font=("Helvetica", 18), height=35, width=380, placeholder_text="Confirm Password")
        self.confirm_password_entry.place(x=60, y=195)
        # Buttons to toggle password visibility
        self.show_password_img1 = CTkImage(Image.open("show_password_img.png"), size=(20, 20))
        self.hide_password_img1 = CTkImage(Image.open("hide_password_img.png"), size=(20, 20))
        self.toggle_password_button1 = CTkButton(master=root, image=self.show_password_img1, text="", width=20, height=30, command=self.toggle_password1)
        self.toggle_password_button1.place(x=450, y=145)
        self.show_password_img2 = CTkImage(Image.open("show_password_img.png"), size=(20, 20))
        self.hide_password_img2 = CTkImage(Image.open("hide_password_img.png"), size=(20, 20))
        self.toggle_password_button2 = CTkButton(master=root, image=self.show_password_img2, text="", width=20, height=30, command=self.toggle_password2)
        self.toggle_password_button2.place(x=450, y=195)
        # Dropdowns for selecting date of birth
        self.months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
        self.days = [f"{i:02d}" for i in range(1, 32)]
        self.years = list(range(1900, date.today().year))[::-1]
        self.selected_day = tk.StringVar()
        self.selected_month = tk.StringVar()
        self.selected_year = tk.StringVar()
        # Month dropdown
        self.month_label = ttk.Label(root, text="Month:")
        self.month_label.place(x=40, y=250)
        self.month_combobox = ttk.Combobox(root,state="readonly", textvariable=self.selected_month, values=self.months)
        self.month_combobox.place(x=40, y=275)
        self.month_combobox.set(self.months[0])
        self.month_combobox.bind("<<ComboboxSelected>>", self.update_days)
        # Day dropdown
        self.day_label = ttk.Label(root, text="Day:")
        self.day_label.place(x=190, y=250)
        self.day_combobox = ttk.Combobox(root,state="readonly", textvariable=self.selected_day, values=self.days)
        self.day_combobox.place(x=190, y=275)
        self.day_combobox.set(self.days[0])
        # Year dropdown
        self.year_label = ttk.Label(root, text="Year:")
        self.year_label.place(x= 340, y=250)
        self.year_combobox = ttk.Combobox(root, state="readonly", textvariable=self.selected_year, values=self.years)
        self.year_combobox.place(x=340,y=275)
        self.year_combobox.set(self.years[0])
        # Label for days available for work
        self.days_not_working_label = CTkLabel(master=root, text="The days you can work", font=("Helvetica", 20), text_color = "#FFFFFF")
        self.days_not_working_label.place(x=50, y=320)
        # Switches for each day of the week
        self.switch_var1 = IntVar(value=1)
        self.monday_switch = CTkSwitch(master=root, text="Monday", command=self.switcher1, variable=self.switch_var1, onvalue=1, offvalue=0)
        self.monday_switch.place(x=50, y=360)
        self.switch_var2 = IntVar(value=1)
        self.tuesday_switch = CTkSwitch(master=root, text="Tuesday", command=self.switcher2, variable=self.switch_var2, onvalue=1, offvalue=0)
        self.tuesday_switch.place(x=150, y=360)
        self.switch_var3 = IntVar(value=1)
        self.wednesday_switch = CTkSwitch(master=root, text="Wednesday", command=self.switcher3, variable=self.switch_var3, onvalue=1, offvalue=0)
        self.wednesday_switch.place(x=250, y=360)
        self.switch_var4 = IntVar(value=1)
        self.thursday_switch = CTkSwitch(master=root, text="Thursday", command=self.switcher4, variable=self.switch_var4, onvalue=1, offvalue=0)
        self.thursday_switch.place(x=370, y=360)
        self.switch_var5 = IntVar(value=1)
        self.friday_switch = CTkSwitch(master=root, text="Friday", command=self.switcher5, variable=self.switch_var5, onvalue=1, offvalue=0)
        self.friday_switch.place(x=50, y=390)
        self.switch_var6 = IntVar(value=1)
        self.saturday_switch = CTkSwitch(master=root, text="Saturday", command=self.switcher6, variable=self.switch_var6, onvalue=1, offvalue=0)
        self.saturday_switch.place(x=150, y=390)
        self.switch_var7 = IntVar(value=1)
        self.sunday_switch = CTkSwitch(master=root, text="Sunday", command=self.switcher7, variable=self.switch_var7, onvalue=1, offvalue=0)
        self.sunday_switch.place(x=250, y=390)
        # Contract hours per week
        self.contract_hours_label = CTkLabel(master=root, text=("Contract minimum hours per week"), font=("Helvetica", 20))
        self.contract_hours_label.place(x=50, y=430)
        self.contract_hours_val = ["8", "16", "24", "32", "40"]
        self.contract_hours_combobox = CTkComboBox(master=root, values=self.contract_hours_val)
        self.contract_hours_combobox.place(x=50, y=465)
        # Primary store selection
        self.current_store_label = CTkLabel(master=root, text="The store you will primarily work at", font=("Helvetica", 20))
        self.current_store_label.place(x=50, y=505)
        self.current_store_val = ["South Kensington", "Battersea", "Hampstead", "East Dulwich", "Chiswick", "Stoke Newington", "Wimbledon"]
        self.current_store_combobox = CTkComboBox(master=root, values=self.current_store_val, font=("Helvetica", 12))
        self.current_store_combobox.place(x=50, y=540)
        # Staff position selection
        self.staff_position_label = CTkLabel(master=root, text="Your current staff position", font=("Helvetica", 20))
        self.staff_position_label.place(x=50, y=580)
        self.staff_position_val = ["Sales Assistant", "Gelato Chef", "Assistant Manager", "Store Manager", "Rota Manager"]
        self.staff_position_combobox = CTkComboBox(master=root, values=self.staff_position_val, font=("Helvetica", 12))
        self.staff_position_combobox.place(x=50, y=615)
        # Button to create the account
        self.create_account_btn = CTkButton(master=root, text="Create Account", fg_color="#0078d7", font=("Helvetica", 16), width=220,
                                            height=40,command=lambda: [self.check_account_info()])
        self.create_account_btn.place(x=140, y=745)
        # Back button
        self.back_button = CTkButton(master=root, text="Back" , font=("Helvetica", 16), width = 220, height = 40, fg_color="#0078d7", hover_color="#DD1D00",
                                     command=lambda: [root.destroy(), self.quit_button()]) 
        self.back_button.place(x=140, y= 795)
        
    def update_days(self, event):
        # Get the selected month from the combobox
        selected_month = self.selected_month.get()
        # Default number of days in a month
        days_in_month = 31
        # Update number of days based on selected month
        if selected_month in ["April", "June", "September", "November"]:
            days_in_month = 30
        elif selected_month == "February":
            # For February, check for leap year
            selected_year = int(self.selected_year.get())
            if (selected_year % 4 == 0 and selected_year % 100 != 0) or (selected_year % 400 == 0):
                days_in_month = 29
            else:
                days_in_month = 28
        # Generate list of days in the month
        self.days = list(range(1, days_in_month + 1))
        # Update the values in the day combobox
        self.day_combobox['values'] = self.days
        # Set the default day value
        self.day_combobox.set(self.days[0])

    def quit_button(self):
        # Decrement the global counter by 1
        global counter1
        counter1 = counter1 - 1
        # Restore visibility of the root window
        root.deiconify()

    # Methods for handling switch events for each day of the week
    def switcher1(self):
        pass

    def switcher2(self):
        pass

    def switcher3(self):
        pass

    def switcher4(self):
        pass

    def switcher5(self):
        pass
        
    def switcher6(self):
        pass
                
    def switcher7(self):
        pass

    def toggle_password1(self):
        # Toggle visibility of password in password entry field
        if self.password_entry.cget("show") == "":
            self.toggle_password_button1.configure(image=self.show_password_img1)
            self.password_entry.configure(show="*")
        else:
            self.toggle_password_button1.configure(image=self.hide_password_img1)
            self.password_entry.configure(show="")

    def toggle_password2(self):
        # Toggle visibility of password in confirm password entry field
        if self.confirm_password_entry.cget("show") == "":
            self.toggle_password_button2.configure(image=self.show_password_img2)
            self.confirm_password_entry.configure(show="*")
        else:
            self.toggle_password_button2.configure(image=self.hide_password_img2)
            self.confirm_password_entry.configure(show="")
            
    def check_account_info(self):
        # Retrieve user input from entry fields
        self.fullname = self.fullname_entry.get()
        self.password = self.password_entry.get()
        self.confirm_password = self.confirm_password_entry.get()
        self.email_address = self.email_entry.get()
        # Check if fullname contains a space
        if self.fullname and " " in self.fullname:
            # Check if fullname contains any digits
            if not re.search(r'\d', self.fullname):
                # Check if email address is provided
                if self.email_address:
                    # Check if email address has correct format
                    if "@" in self.email_address and ".com" in self.email_address:
                        # Check if password and confirm password are provided
                        if self.password and self.confirm_password:
                            # Check if password and confirm password match
                            if self.password == self.confirm_password:
                                # Check if password length is at least 6 characters
                                if len(self.password) >= 6:
                                    # Hash the password
                                    self.hashed_password = bcrypt.hashpw((self.password.encode("utf-8")), bcrypt.gensalt())
                                    # Extract day, month, and year from date dropdowns
                                    self.day = str(self.day_combobox.get())
                                    self.month = str(self.month_combobox.get())
                                    self.year = str(self.year_combobox.get())
                                    # Map month name to month number
                                    self.months_dict = {"January": "01",
                                        "February": "02",
                                        "March": "03",
                                        "April": "04",
                                        "May": "05",
                                        "June": "06",
                                        "July": "07",
                                        "August": "08",
                                        "September": "09",
                                        "October": "10",
                                        "November": "11",
                                        "December": "12"}
                                    self.month = self.months_dict[self.month]
                                    # Concatenate day, month, and year to form birthday
                                    self.birthday = self.day+"/"+self.month+"/"+self.year
                                    # Determine if user is available to work on each day of the week
                                    self.monday_working = 1 if self.switch_var1.get() == 1 else 0
                                    self.tuesday_working = 1 if self.switch_var2.get() == 1 else 0
                                    self.wednesday_working = 1 if self.switch_var3.get() == 1 else 0
                                    self.thursday_working = 1 if self.switch_var4.get() == 1 else 0
                                    self.friday_working = 1 if self.switch_var5.get() == 1 else 0
                                    self.saturday_working = 1 if self.switch_var6.get() == 1 else 0
                                    self.sunday_working = 1 if self.switch_var7.get() == 1 else 0
                                    # Check if at least one day is marked as working
                                    if any([self.monday_working, self.tuesday_working, self.wednesday_working, self.thursday_working, self.friday_working, self.saturday_working, self.sunday_working]):
                                        try:
                                            int(self.contract_hours_combobox.get())
                                        except:
                                            # Display error if invalid contract hours entered
                                            hours_error = tk.messagebox.showerror(title="Error", message="Invalid contract hours entered")
                                        else:
                                            # Check if contract hours are within valid range
                                            r = range(1,52)
                                            self.contract_hours = int(self.contract_hours_combobox.get())
                                            if self.contract_hours in r:
                                                # Map current store to closest store
                                                self.closest_store_dic = {"South Kensington" : "Chiswick", "Battersea" : "East Dulwich", "Hampstead" : "Stoke Newington",
                                                                          "East Dulwich" : "Battersea", "Chiswick" : "South Kensington", "Stoke Newington" : "Hampstead", "Wimbledon" : "Battersea"}
                                                self.current_store = self.current_store_combobox.get()
                                                # Check if current store is valid
                                                if self.current_store in self.closest_store_dic:
                                                    self.closest_store = self.closest_store_dic.get(self.current_store)
                                                    self.position = self.staff_position_combobox.get()
                                                    # Check if staff position is valid
                                                    if self.position in self.staff_position_val:
                                                        # Determine if user is admin or not
                                                        if self.position == "Sales Assistant":
                                                            self.admin = 0
                                                        elif self.position == "Gelato Chef":
                                                            self.admin = 1
                                                        elif self.position == "Assistant Manager":
                                                            self.admin = 2
                                                        elif self.position == "Store Manager":
                                                            self.admin = 3
                                                        elif self.position == "Rota Manager":
                                                            self.admin = 4
                                                            
                                                        # Save account information
                                                        self.save_account_info()
                                                    else:
                                                        # Display error if invalid staff position entered
                                                        position_error = tk.messagebox.showerror(title="Error", message="Invalid staff position")
                                                else:
                                                    # Display error if invalid store entered
                                                    store_error = tk.messagebox.showerror(title="Error", message="Invalid Store entered")
                                            else:
                                                # Display error if invalid contract hours entered
                                                hours_error = tk.messagebox.showerror(title="Error", message="Invalid contract hours entered")
                                    else:
                                        # Display error if no days marked as working
                                        working_error = tk.messagebox.showerror(title="Error", message="Invalid days working entered")
                                else:
                                    # Display error if password length is less than 6 characters
                                    len_password_error = tk.messagebox.showerror(title="Error", message="Use 6 characters or more for your password")
                            else:
                                # Display error if passwords do not match
                                password_error = tk.messagebox.showerror(title="Error", message="Passwords do not match")
                        else:
                            # Display error if password or confirm password fields are empty
                            empty_password_error = tk.messagebox.showerror(title="Error", message="Fill in the Password fields ")
                    else:
                        # Display error if invalid email address format
                        email_error = tk.messagebox.showerror(title="Error", message="Invalid Email Address")
                else:
                    # Display error if email address field is empty
                    empty_email_error = tk.messagebox.showerror(title="Error", message="Fill in the Email Address field")
            else:
                # Display error if fullname contains digits
                name_error = tk.messagebox.showerror(title="Error", message="Numbers are not allowed in Full Name")
        else:
            # Display error if fullname field is empty
            empty_name_error = tk.messagebox.showerror(title="Error", message="Fill in the Full Name field")

    # Define a method to save account information
    def save_account_info(self):
        # Initialize userid variable to None
        userid = None
        # Assign fullname attribute to variable
        fullname = self.fullname
        # Assign email_address attribute to variable
        email = self.email_address
        # Assign hashed_password attribute to variable
        password = self.hashed_password
        # Assign birthday attribute to variable
        birthday = self.birthday
        # Assign monday_working attribute to variable
        monday_working = self.monday_working
        # Assign tuesday_working attribute to variable
        tuesday_working = self.tuesday_working
        # Assign wednesday_working attribute to variable
        wednesday_working = self.wednesday_working
        # Assign thursday_working attribute to variable
        thursday_working = self.thursday_working
        # Assign friday_working attribute to variable
        friday_working = self.friday_working
        # Assign saturday_working attribute to variable
        saturday_working = self.saturday_working
        # Assign sunday_working attribute to variable
        sunday_working = self.sunday_working
        # Assign contract_hours attribute to variable
        contract_hours = self.contract_hours
        # Assign current_store attribute to variable
        primary_store = self.current_store
        # Assign closest_store attribute to variable
        closest_store = self.closest_store
        # Assign admin attribute to variable
        admin = self.admin
        # Connect to the SQLite database
        conn = sqlite3.connect('Rota_Maker.db')
        # Create a cursor object
        cursor = conn.cursor()
        # Execute SQL query to retrieve StoreID for primary_store
        cursor.execute("SELECT StoreID FROM Stores WHERE Store_Name=(?)", (primary_store,))
        # Fetch the primary_store_key
        primary_store_key = cursor.fetchall()[0][0]
        # Execute SQL query to retrieve StoreID for closest_store
        cursor.execute("SELECT StoreID FROM Stores WHERE Store_Name=(?)", (closest_store,))
        # Fetch the closest_store_key
        closest_store_key = cursor.fetchall()[0][0]
        # Execute SQL query to check if email already exists in Account table
        cursor.execute("SELECT Email FROM Account WHERE Email=(?)", (email,))
        # Fetch the result of email check
        checkemail = cursor.fetchall()
        # Execute SQL query to retrieve all names from Account table
        checknames = cursor.execute("SELECT Name FROM Account").fetchall()
        # Initialize a list to store first names
        firstnames = []
        # Iterate over checknames and extract first names
        for i in checknames:
            # Split the string and extract first name
            first = str(i).split()[0]
            # Append first name to firstnames list
            firstnames.append(first)
        # Iterate over firstnames list
        for i in firstnames:
            # Check if first name matches with any existing names
            if fullname.split()[0] in i:
                # Show error message if there is a duplicate name
                tk.messagebox.showerror("Error", "There is already an account with this name, put a nickname or shortened name to avoid duplicates")
                # Return from the function
                return
        # Check if email does not already exist in Account table
        if checkemail == []:
            # Insert account information into Account table
            cursor.execute("INSERT INTO Account VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                           (userid, fullname, email, password, birthday, monday_working, tuesday_working,
                            wednesday_working, thursday_working, friday_working, saturday_working, sunday_working,
                            contract_hours, primary_store_key, closest_store_key, admin))
            # Commit changes to the database
            conn.commit()
            # Close the database connection
            conn.close()
            # Show message box indicating successful account creation
            data_inserted = tk.messagebox.showinfo(message="Account Created", command=self.complete_button())
        else:
            # Show error message if account already exists with the email
            tk.messagebox.showerror(title="Error", message="Account already exists with this email")
            # Close the database connection
            conn.close()

    # Define a method to handle completion of a button action
    def complete_button(self):
        # Destroy the root window
        self.root.destroy()
        # Access the global counter1 variable and decrement its value by 1
        global counter1
        counter1 = counter1 - 1
        # Redisplay the root window
        root.deiconify()

# Base class with common functionalities
class default_window:
    # Constructor method to initialize the class
    def __init__(self, root, login_email):
        # Initialize attributes
        self.root = root
        self.login_email = login_email
        # Connect to the database
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        # Retrieve data from the Account table based on login_email
        cursor.execute("SELECT admin, name, Primary_Store FROM Account WHERE Email=(?)", (self.login_email,))
        data = cursor.fetchall()
        conn.close()
        # Assign retrieved data to attributes
        self.admin = int(data[0][0])
        self.name = data[0][1]
        self.store = data[0][2]
        self.name_caps = self.name.split()
        x = self.name_caps[0].capitalize()
        y = self.name_caps[-1].capitalize()
        self.name_caps = x + " " + y
        var = StringVar(value=self.name_caps)
        # Common initialization code for all admin levels
        self.root.title("Rota Maker")
        self.root.geometry("675x575+645+150")
        self.root.resizable(False, False)
        self.top_middle_frame = CTkFrame(master=root, width=635, height=70, fg_color="#6A5ACD")
        self.top_middle_frame.place(x=20, y=20)
        self.welcome_label = CTkLabel(master=self.top_middle_frame, text=f"Welcome, {var.get()}!", font=("Arial", 32), text_color="#FFFFFF")
        self.welcome_label.place(relx=0.5, y=35, anchor="center")
        # Buttons for common functionalities
        self.next_weeks_button = CTkButton(master=root, width = 200, height = 60, fg_color="#6A5ACD", corner_radius=10, text="Next Week's Rotas",
                                           font=("Arial", 20), text_color="#FFFFFF", hover_color="#3F318B", command=lambda:[self.next_week()])
        self.next_weeks_button.place(x=237.5, y=140)
        Tooltip(self.next_weeks_button, "View the rotas for next week (for all stores)")
        self.log_out_button = CTkButton(master=root, width=120, height=60, fg_color="#E51E1E", corner_radius=5, text="Log out", font=("Arial", 20),
                                        text_color="#FFFFFF", hover_color="#E51E1F", command=self.log_out)
        self.log_out_button.place(x=10, y=505)
        self.change_account_details_button = CTkButton(master=root, width=120, height=60, fg_color="#85B743", corner_radius=5, text="Account Settings",
                                                       font=("Arial", 20), text_color="#FFFFFF", hover_color="#6B8845", command=self.change_account_details)
        self.change_account_details_button.place(x=500, y=505)

    # Method to view rotas
    def view_rotas(self):
        global view_counter
        if view_counter == 0:
            view_counter += 1
            root1 = CTkToplevel(self.root)
            View_Rotas(root1)

    # Method to handle next week's rotas
    def next_week(self):
        global next_week
        if next_week == 0:
            next_week += 1
            root1 = CTkToplevel(self.root)
            Next_Week(root1)

    # Method to create a rota
    def create_rota(self):
        global create_counter
        if create_counter == 0:
            create_counter += 1
            root1 = CTkToplevel(self.root)
            Create_Rotas(root1, self.name, self.admin)

    # Method to publish rotas
    def publish_rotas(self):
        global publish_counter
        if publish_counter == 0:
            publish_counter += 1
            root1 = CTkToplevel(self.root)
            Publish_Rotas(root1)

    # Method to view worker information
    def workers_info(self):
        global info_counter
        if info_counter == 0:
            info_counter += 1
            root1 = CTkToplevel(self.root)
            Workers_Info(root1)
        
    # Method to handle log out functionality
    def log_out(self):
        dialogue = False
        if dialogue == False:
            dialogue = True
            answer = tk.messagebox.askyesno(title="Confirmation", message="Are you sure you want to log out?")
            if answer:
                self.root.destroy()
                variables_to_check = ['view_counter', 'create_counter', 'publish_counter', 'info_counter', 'account_settings_counter', 'next_week']
                for var_name in variables_to_check:
                    if globals().get(var_name) == 1:
                        globals()[var_name] = 0
                self.log_out_successful()
            else:
                dialogue = False
        else:
            pass

    # Method to handle successful log out
    def log_out_successful(self):
        global counter1
        counter1 = 0
        root.deiconify()

    # Method to handle account details change
    def change_account_details(self):
        global account_settings_counter
        if account_settings_counter == 0:
            account_settings_counter += 1
            root1 = CTkToplevel(self.root)
            Account_Settings(root1, self.login_email)

# Class for employee window
class employee_window(default_window):
    # Constructor method
    def __init__(self, root, login_email):
        # Call parent class constructor
        super().__init__(root, login_email)
        # Connect to database
        conn = sqlite3.connect("Rota_Maker.db")
        cursor = conn.cursor()
        # Retrieve store name
        store = cursor.execute("SELECT Store_Name FROM Stores WHERE StoreID =(?)", (self.store,)).fetchall()[0][0]
        conn.close()
        var1 = StringVar(value=store)
        # Determine position based on admin level
        if self.admin == 1:
            pos = "Gelato Chef"
        else:
            pos = "Sales Assistant"
        # Label to display position and store
        self.you_are_label = CTkLabel(root, text=f"{pos} at {var1.get()}", font=("Arial", 16))
        self.you_are_label.place(x=10, y=100)
        # Button to view all completed rotas
        self.view_rotas_button = CTkButton(master=root, width = 200, height = 60, fg_color="#6A5ACD", corner_radius=10, text="All Completed Rotas",
                                           font=("Arial", 20), text_color="#FFFFFF", hover_color="#3F318B", command=lambda:[self.view_rotas()])
        self.view_rotas_button.place(x=237.5, y=230)
        Tooltip(self.view_rotas_button, "View the previously made rotas")

# Class for manager window
class manager_window(default_window):
    # Constructor method
    def __init__(self, root, login_email):
        # Call parent class constructor
        super().__init__(root, login_email)
        # Connect to database
        conn = sqlite3.connect("Rota_Maker.db")
        cursor = conn.cursor()
        # Retrieve store name
        store = cursor.execute("SELECT Store_Name FROM Stores WHERE StoreID =(?)", (self.store,)).fetchall()[0][0]
        conn.close()
        var1 = StringVar(value=store)
        # Determine position based on admin level
        if self.admin == 3:
            pos = "Store Manager"
            # Label to display position and store
            self.you_are_label = CTkLabel(root, text=f"{pos} of {var1.get()}", font=("Arial", 16))
            self.you_are_label.place(x=10, y=100)
        elif self.admin == 2:
            pos = "Assistant Manager"
            # Label to display position and store
            self.you_are_label = CTkLabel(root, text=f"{pos} of {var1.get()}", font=("Arial", 16))
            self.you_are_label.place(x=10, y=100)
        elif self.admin == 4:
            pos = "Rota Manager"
            # Label to display position and store
            self.you_are_label = CTkLabel(root, text=f"{pos}", font=("Arial", 16))
            self.you_are_label.place(x=10, y=100)

        # Button to create rota
        self.create_rotas_button = CTkButton(master=root, width=200, height=60, fg_color="#6A5ACD", corner_radius=10, text="Create Rota",
                                             font=("Arial", 20), text_color="#FFFFFF", hover_color="#3F318B", command=self.create_rota)
        self.create_rotas_button.place(x=237.5, y=230)
        Tooltip(self.create_rotas_button, "Create a new rota")
        # Button to view worker information
        self.workers_info_button = CTkButton(master=root, width=200, height=60, fg_color="#6A5ACD", corner_radius=10, text="Worker Info",
                                             font=("Arial", 20), text_color="#FFFFFF", hover_color="#3F318B", command=self.workers_info)
        self.workers_info_button.place(x=237.5, y=320)
        Tooltip(self.workers_info_button, """View information about
the workers in each store""")
        # Button to view all completed rotas
        self.view_rotas_button = CTkButton(master=root, width = 200, height = 60, fg_color="#6A5ACD", corner_radius=10, text="All Completed Rotas",
                                           font=("Arial", 20), text_color="#FFFFFF", hover_color="#3F318B", command=lambda:[self.view_rotas()])
        self.view_rotas_button.place(x=237.5, y=410)
        Tooltip(self.view_rotas_button, "View the previously made rotas")

# Class for creating tooltips
class Tooltip:
    # Constructor method
    def __init__(self, widget, text, delay=1100):
        # Widget to which tooltip is attached
        self.widget = widget
        # Text to be displayed in the tooltip
        self.text = text
        # Delay before showing the tooltip
        self.delay = delay
        # Initialize tooltip variable
        self.tooltip = None
        # Bind Enter event to on_enter method
        self.widget.bind("<Enter>", self.on_enter)
        # Bind Leave event to hide_tooltip method
        self.widget.bind("<Leave>", self.hide_tooltip)

    # Method to handle mouse entering the widget
    def on_enter(self, event=None):
        # Schedule showing the tooltip
        self.tooltip = self.widget.after(self.delay, lambda: self.show_tooltip(event))

    # Method to handle mouse leaving the widget
    def on_leave(self, event=None):
        if self.tooltip:
            # Cancel scheduled tooltip
            self.widget.after_cancel(self.tooltip)
            # Reset tooltip variable
            self.tooltip = None
            # Hide the tooltip
            self.hide_tooltip()

    # Method to show the tooltip
    def show_tooltip(self, event=None):
        if self.tooltip:
            # Calculate tooltip position
            x, y, _, _ = self.widget.bbox("insert")
            x += self.widget.winfo_rootx() + 204
            y += self.widget.winfo_rooty() + 10
            # Create a new toplevel window for the tooltip
            self.tooltip = CTkToplevel(self.widget)
            self.tooltip.wm_overrideredirect(True)
            self.tooltip.wm_geometry(f"+{x}+{y}")
            # Label to display tooltip text
            label = tk.Label(self.tooltip, text=self.text, justify="left", background="#D2D2D2", relief="solid", borderwidth=1)
            label.pack(ipadx=1)
            # Bind Leave event to on_leave method
            self.tooltip.bind("<Leave>", self.on_leave)

    # Method to hide the tooltip
    def hide_tooltip(self, event=None):
        try:
            if self.tooltip:
                # Cancel scheduled tooltip
                self.widget.after_cancel(self.tooltip)
                # Destroy the tooltip window
                self.tooltip.destroy()
                # Reset tooltip variable
                self.tooltip = None
        except:
            # Ignore any errors during tooltip hiding
            pass

# Initialize the View_Rotas class, which handles the view of rotas
class View_Rotas:
    # Constructor method for initializing the class
    def __init__(self, root):
        # Set the root window for the view
        self.root = root
        # Set the title of the root window
        self.root.title("View Rotas")
        # Set the dimensions and position of the root window
        self.root.geometry("400x400+1420+100")
        # Make the root window non-resizable
        self.root.resizable(False, False)
        # Set the background color of the root window
        self.root.configure(bg="#272727")
        # Function to handle window closing
        def on_closing():
            # Access global variable view_counter
            global view_counter
            # Decrement the counter when the window is closed
            view_counter -= 1
            # Destroy the root window
            self.root.destroy()
        # Bind the closing event of the root window to the on_closing function
        self.root.protocol("WM_DELETE_WINDOW", on_closing)
        # Label for date selection
        self.date_label = tk.Label(self.root, text="Select the week starting: ", font=("Arial", 13))
        self.date_label.place(x=90, y=30)
        # Retrieve last year's Mondays and store them
        self.monday_dates = self.get_last_year_mondays()
        self.selected_monday = tk.StringVar()
        # Dropdown for selecting Monday
        self.monday_combobox = ttk.Combobox(root, state="readonly", width=15, font=("Arial", 11), textvariable=self.selected_monday, values=self.monday_dates)
        self.monday_combobox.place(x=110, y=60)
        self.monday_combobox.set(self.monday_dates[0])
        # Label for store selection
        self.store_label = tk.Label(self.root, text="Select a store:", font=("Arial", 13))
        self.store_label.place(x=90, y=100)
        # Variable to store selected store
        self.store_var = tk.StringVar(self.root)
        # Dropdown for selecting store
        self.store_dropdown = ttk.Combobox(self.root, font=("Arial", 11), textvariable=self.store_var)
        self.store_dropdown['values'] = ['South Kensington', 'Battersea', 'Hampstead', 'East Dulwich', 'Chiswick', 'Stoke Newington', 'Wimbledon']
        self.store_dropdown.place(x=110, y=130)
        self.store_dropdown.set("South Kensington")
        self.store_dropdown.configure(state="readonly")
        # Button to save selection
        self.save_button = CTkButton(self.root, text="Save", command=self.save_selection)
        self.save_button.place(x=110, y=160)
        # Start the Tkinter event loop
        self.root.mainloop()

    # Method to get last year's Mondays
    def get_last_year_mondays(self):
        # Get today's date
        today = datetime.datetime.today()
        # Calculate last year
        last_year = today.year - 1
        # Initialize list to store Mondays
        mondays = []
        # Start from the current date and move back to the last Monday
        current_date = today - timedelta(days=today.weekday())
        while current_date.year >= last_year:
            # Check if the current date is a Monday
            if current_date.weekday() == 0:
                mondays.append(current_date.strftime("%d/%m/%Y"))
            current_date -= timedelta(days=1)
        return mondays
    
    # Method to save user selection
    def save_selection(self):
        # Get selected date and store
        selected_date = self.monday_combobox.get()
        self.selected_store = self.store_dropdown.get()
        # Check if both date and store are selected
        if not selected_date or not self.selected_store:
            # Show error message if either date or store is not selected
            tk.messagebox.showerror("Error", "Please select a date and a store.")
            return
        # Format store name
        formatted_store_name = self.selected_store.replace(" ", "_")
        # Format date
        formatted_date = datetime.datetime.strptime(selected_date, "%d/%m/%Y").strftime("%d_%m_%Y")
        # Generate table name
        self.table_name = f"finished_rota_{formatted_store_name}_{formatted_date}"
        # Display data from the database
        self.display_data_from_database()

    # Method to display data from the database
    def display_data_from_database(self):
        # Connect to SQLite database
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        try:
            # Check if the table exists
            cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name=?", (self.table_name,))
            existing_table = cursor.fetchone()
            if existing_table:
                # If table exists, retrieve data from it
                select_data_query = f"SELECT Staff_Name, Monday, Tuesday, Wednesday, Thursday, Friday, Saturday, Sunday, Total_Hrs FROM {self.table_name}"
                data_from_db = cursor.execute(select_data_query).fetchall()
            else:
                # If table doesn't exist, show error message
                tk.messagebox.showerror(title="Error", message="There is no rota made for this week")
                return
        except:
            pass
        # Create a popup window to display data
        popup = tk.Toplevel(self.root)
        popup.title('Data Display')
        popup.resizable(0, 0)
        # Get store names from database
        storesdict = ["South Kensington", "Battersea", "Hampstead", "East Dulwich", "Chiswick", "Stoke Newington", "Wimbledon"]
        temp_number = int(storesdict.index(self.selected_store) + 1)
        temp_names = cursor.execute("SELECT Name FROM Account WHERE Primary_Store = (?)", (temp_number,)).fetchall()
        temp_names_len = len(temp_names)
        h = ((temp_names_len + 1) * 3) + 1
        self.height = (h * 22)
        if self.height > 900:
            self.height = 900
        popup.geometry(f"1380x{self.height}")
        # Create a Treeview widget
        tree = ttk.Treeview(popup, height=h)
        # Define columns
        tree["columns"] = ("Staff Name", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday", "Total Hrs")
        # Format columns
        tree.column("#0", width=0, stretch=tk.NO)  # Hide the first column
        for col in tree["columns"]:
            if col == "Staff Name":
                tree.column(col, width=178, minwidth=178, anchor=tk.CENTER)
                tree.heading(col, text=col)
            elif col in ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"):
                tree.column(col, width=156, minwidth=156, anchor=tk.CENTER, stretch=tk.NO)
                tree.heading(col, text=col)
            elif col == "Total Hrs":
                tree.column(col, width=90, minwidth=90, anchor=tk.CENTER, stretch=tk.NO)
        # Configure tags for styling
        tree.tag_configure('bold', font=('Arial', 10, 'bold'))
        tree.tag_configure('font_size', font=('Arial', 9))
        # Insert data into Treeview
        for index, row in enumerate(data_from_db):
            # Apply 'bold' tag to the staff name row and the last row (budget hours)
            tags = ("bold",) if index % 3 == 0 or index == len(data_from_db) - 2 else ("font_size",)
            tree.insert("", "end", values=row, tags=tags)
        # Resize popup window based on data
        x = (len(data_from_db) * 22)
        popup.geometry(f"1380x{x}")
        # Add Treeview to popup window
        tree.pack(fill="both", expand=True)
        # Close the database connection
        conn.close()

# Define a class named Node
class Node:
    # Define a constructor for the Node class, initializing data and next
    def __init__(self, data):
        # Assign the data attribute of the node
        self.data = data
        # Initialize the next attribute of the node to None
        self.next = None

# Define a class named LinkedList
class LinkedList:
    # Define a constructor for the LinkedList class, initializing head to None
    def __init__(self):
        # Initialize the head attribute of the linked list to None
        self.head = None

    # Define a method to append data to the linked list
    def append(self, data):
        # Create a new node with the provided data
        new_node = Node(data)
        # Check if the linked list is empty
        if not self.head:
            # If empty, set the new node as the head
            self.head = new_node
        else:
            # Start traversal from the head
            current = self.head
            # Traverse to the end of the linked list
            while current.next:
                # Move to the next node
                current = current.next
            # Set the next of the current node to the new node
            current.next = new_node

    # Define a method to remove a node with a given value from the linked list
    def remove(self, value):
        # Start traversal from the head
        current = self.head
        # Initialize a pointer to keep track of the previous node
        previous = None
        # Initialize a flag to indicate if the value is found
        found = False
        # Traverse the list to find the node containing the value
        while current and not found:
            # Iterate until the end of the list or value is found
            if current.data == value:
                # Check if the current node contains the value
                found = True
            else:
                # Update the previous node pointer
                previous = current
                # Move to the next node
                current = current.next
        # If the value was found, remove the node
        if found:
            # If the node to remove is not the head
            if previous:
                # Update the next of the previous node
                previous.next = current.next
            else:
                # If the node to remove is the head, update the head
                self.head = current.next
            # Delete the node
            del current
        else:
            # If the value is not found, do nothing
            return

class Create_Rotas:
    def __init__(self, root, name, admin):
        # Counter for saved dates
        self.save_date_count = 0  
        # Reference to the Tkinter root window
        self.root = root  
        # Set the title of the window
        self.root.title("Create Rota")  
        # Set the window dimensions and position
        self.root.geometry("400x750+50+50")  
        # Make the window non-resizable
        self.root.resizable(False, False)  
        # Set the background color of the window
        self.root.configure(bg="#272727")  
        # User name
        self.name = name  
        # Admin status of the user
        self.admin = admin  
        # Function to handle window closing
        def on_closing():
            # Access the global counter variable
            global create_counter  
            # Decrement the counter
            create_counter -= 1  
            # Destroy the root window when closing
            self.root.destroy()  
        # Register the on_closing function for window closing event
        self.root.protocol("WM_DELETE_WINDOW", on_closing)  
        # Get today's date and calculate the date of last Monday
        today = datetime.datetime.today()
        last_monday = today - timedelta(days=today.weekday()) 
        # Generate dates for the next 3 weeks
        dates = []
        for i in range(1, 4):
            date = last_monday + timedelta(weeks=i)
            dates.append(date)
        dates = [date.strftime('%d/%m/%Y') for date in dates]
        self.dates = dates
        # Connect to the database and retrieve the primary store of the user
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        # Fetch the primary store name associated with the user
        self.their_store = cursor.execute("SELECT Store_Name FROM Account a, Stores b WHERE a.Primary_Store = b.StoreID AND Name = (?)", (self.name,)).fetchall()[0]
        conn.close()  # Close the database connection
        # Extract the store name from the fetched result
        for i in self.their_store:
            self.their_store = i
        # GUI elements creation
        # Label for store
        self.store_label = CTkLabel(master=root, text="Store:", font=("Arial", 16))  
        # Place the store label
        self.store_label.place(x=18, y=23)  
        # Available stores
        self.stores = ["South Kensington", "Battersea", "Hampstead", "East Dulwich", "Chiswick", "Stoke Newington", "Wimbledon"]  
        # Dropdown for stores
        self.store_dropdown = CTkComboBox(master=root, values=self.stores, font=("Arial", 16), width=175, state="readonly")  
        # Set default store
        self.store_dropdown.set(self.stores[0])  
        # Place the store dropdown
        self.store_dropdown.place(x=15, y=50)  
        # Button to save store
        self.save_store_button = CTkButton(master=root, text="Save", font=("Arial", 16), fg_color="#0078d7",width=70, command=lambda: [self.save_store_clicked()])  
        # Place the save store button
        self.save_store_button.place(x=15, y=85)  
        # Button to undo store
        self.undo_store_button = CTkButton(master=root, text="Undo", font=("Arial", 16), fg_color="#0078d7", width=70, command=lambda: [self.undo_store_clicked()])  
        # Initially hide undo store button
        self.undo_store_button.place_forget()  
        # Label for week beginning
        self.date_label = CTkLabel(master=root, text="Week beginning:", font=("Arial", 16))  
        # Place the week beginning label
        self.date_label.place(x=208, y=23)  
        # Dropdown for week beginning
        self.date_dropdown = CTkComboBox(master=root, values=self.dates, font=("Arial", 16), state="disabled")  
        # Place the week beginning dropdown
        self.date_dropdown.place(x=205, y=50)  
        # Button to save date
        self.save_date_button = CTkButton(master=root, text="Save", font=("Arial", 16), fg_color="#0078d7",width=70, command=lambda: [self.save_date_clicked()])  
        # Initially hide save date button
        self.save_date_button.place_forget()  
        # Button to undo date
        self.undo_date_button = CTkButton(master=root, text="Undo", font=("Arial", 16), fg_color="#0078d7", width=70, command=lambda: [self.undo_date_clicked()])  
        # Initially hide undo date button
        self.undo_date_button.place_forget()

    def save_store_clicked(self):
        # Connect to the database
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        # Fetch admin status of the user
        admin = cursor.execute("SELECT Admin FROM Account WHERE Name =?", (self.name,)).fetchall()[0][0]
        conn.close()  # Close the database connection
        # Check if the user is authorized to save the store
        if self.store_dropdown.get() == self.their_store or admin == 4:
            # Check if the store dropdown is in readonly state
            if self.store_dropdown.cget("state") == "readonly":
                # Disable the store dropdown and enable the date dropdown
                self.store_dropdown.configure(state="disabled")
                self.date_dropdown.configure(state="readonly")
                # Set default date if none selected
                if not self.date_dropdown.get():
                    self.date_dropdown.set(self.dates[0])
                # Hide the save store button and show the undo store button
                self.save_store_button.place_forget()
                self.undo_store_button.place(x=95,y=85)
                # Get the selected store
                self.store = self.store_dropdown.get()
                # Fetch store number from the database
                conn = sqlite3.connect('Rota_Maker.db')
                cursor = conn.cursor()
                self.store_number = cursor.execute("SELECT StoreID From Stores WHERE Store_Name = (?)", (self.store,)).fetchall()
                conn.close()  # Close the database connection
                # Show the save date button
                self.save_date_button.place(x=205, y=85)
        else:
            # Show error message if user is not authorized
            tk.messagebox.showerror(title="Error", message="You can only make the rota for your store")

    def undo_store_clicked(self):
        # Check if a date is saved, if so, undo it
        if self.save_date_count == 1:
            self.undo_date_clicked()
        # Set the store dropdown back to readonly state
        self.store_dropdown.configure(state="readonly")
        # Disable the date dropdown
        self.date_dropdown.configure(state="disabled")
        # Hide the undo store button
        self.undo_store_button.place_forget()
        # Show the save store button
        self.save_store_button.place(x=15, y=85)
        # Hide the save date button
        self.save_date_button.place_forget()

    def save_date_clicked(self):
        # Set the save date count to 1 to indicate a date is saved
        self.save_date_count = 1
        # Get the selected date
        self.date = self.date_dropdown.get()
        # Disable the date dropdown
        self.date_dropdown.configure(state="disabled")
        # Show the undo date button
        self.undo_date_button.place(x=285 , y=85)
        # Hide the save date button
        self.save_date_button.place_forget()
        # Label for staff name
        self.staff_name_label = tk.Label(self.root, text="Staff Name:")
        self.staff_name_label.place(x=10, y=330)
        # Label for day
        self.day_label = tk.Label(self.root, text="Day:")
        self.day_label.place(x=170, y=330)
        # Label for start time
        self.start_time_label = tk.Label(self.root, text="Start Time:")
        self.start_time_label.place(x=10, y=410)
        # Label for end time
        self.end_time_label = tk.Label(self.root, text="End Time:")
        self.end_time_label.place(x=170, y=410)
        # Try to fetch the store number
        try:
            self.store_number1 = self.store_number[0]
        except:
            pass
        else:
            # Connect to the database
            conn = sqlite3.connect('Rota_Maker.db')
            cursor = conn.cursor()
            # Fetch names associated with the store
            self.names_temp = cursor.execute("SELECT Name FROM Account WHERE Primary_Store = (?)", (self.store_number1)).fetchall()
            conn.close()
            # Extract and format names
            names = []
            array = []
            for i in self.names_temp:
                temp = i[0]
                names.append(temp)
            self.first_names = []
            for i in names:
                temp = []
                first_name=i.split()[0].capitalize()
                self.first_names.append(first_name)
        # Combobox for staff names
        self.staff_name_combobox = ttk.Combobox(self.root,state="readonly", values=self.first_names)
        if self.first_names:
            self.staff_name_combobox.set(self.first_names[0])
        else:
            self.staff_name_combobox.set("")
        self.staff_name_combobox.place(x=10, y=360)
        # Days of the week
        self.days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        # Combobox for selecting day
        self.day_combobox = ttk.Combobox(self.root, state="readonly", values=self.days)
        self.day_combobox.set(self.days[0])
        self.day_combobox.place(x=170, y=360)
        # Start times for shifts
        self.start_times = [f"{hour:02d}:{minute:02d}" for hour in range(9, 24) for minute in range(0, 60, 15)]
        # Combobox for selecting start time
        self.start_time_combobox = ttk.Combobox(self.root, state="readonly", values=self.start_times)
        self.start_time_combobox.set(self.start_times[0])
        self.start_time_combobox.place(x=10, y=440)
        # End times for shifts
        self.end_times = [f"{hour:02d}:{minute:02d}" for hour in range(9, 24) for minute in range(0, 60, 15)]
        # Combobox for selecting end time
        self.end_time_combobox = ttk.Combobox(self.root, state="readonly", values=self.end_times)
        self.end_time_combobox.set(self.end_times[0])
        self.end_time_combobox.place(x=170, y=440)
        # Fetch store name for error handling
        storename = None
        try:
            conn = sqlite3.connect('Rota_Maker.db')
            cursor = conn.cursor()
            primid = self.store_number[0][0]
            secid = cursor.execute("SELECT Closest_Store FROM Account WHERE Primary_Store = (?)", (primid,)).fetchall()[0][0]
            self.storename = cursor.execute("SELECT Store_Name FROM Stores WHERE StoreID = (?)", (secid,)).fetchall()[0][0]
            conn.close()
        except:
            # Show error if no one is working at the store
            tk.messagebox.showerror("Error", "There is no one working at this store")
        # Label for new staff member from the same store
        self.new_person_label = tk.Label(self.root, text=f"Staff from {self.storename}:")
        self.new_person_label.place(x=10, y=200)
        # Fetch staff names associated with the store
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        tempnames = cursor.execute("SELECT Name FROM Account WHERE Closest_Store = (?)", (self.store_number1)).fetchall()
        self.staff_names = []
        for i in tempnames:
            name = i[0].capitalize()
            first_name = name.split(" ")[0]
            self.staff_names.append(first_name)
        # Create linked list for staff names
        self.staff_names_linked_list = LinkedList()
        for name in self.staff_names:
            self.staff_names_linked_list.append(name)
        self.other_store = self.staff_names[:]
        tempnames1 = cursor.execute("SELECT Name FROM Account WHERE Primary_Store = (?)", (self.store_number1)).fetchall()
        self.staff_names2 = []
        for i in tempnames1:
            name1 = i[0].capitalize()
            first_name1 = name1.split(" ")[0]
            self.staff_names2.append(first_name1)
        # Combobox for selecting new staff member
        self.new_person_combobox = ttk.Combobox(self.root)
        self.new_person_combobox.place(x=10, y=230)
        if self.staff_names:
            self.populate_combobox(self.staff_names_linked_list)
        else:
            self.new_person_combobox.set("")
        self.new_person_combobox.configure(state="readonly")
        # Button for adding new staff member
        self.add_new_person_button = CTkButton(self.root, text="Add New Staff Member",width=16, command=lambda: self.add_new_person(self.new_person_combobox.get()))
        self.add_new_person_button.place(x=200, y=230)
        # Button for removing staff member from another store
        self.remove_button = CTkButton(self.root, text="Remove Staff from other store", command=self.remove_selected)
        self.remove_button.place(x=10, y=270)
        # Separator between buttons and other UI elements
        self.sep1 = ttk.Separator(self.root, orient='horizontal')
        self.sep1.place(x=0, y=310, relwidth=1.2)
        # Button for adding data
        self.add_button = CTkButton(self.root, text="Add Data",width=16, command=lambda: self.add_data(
            self.staff_name_combobox.get(), self.day_combobox.get(), self.start_time_combobox.get(), self.end_time_combobox.get()))
        self.add_button.place(x=70, y=500)
        # Button for covering shift
        self.cover_shift_button = CTkButton(self.root, text="Cover Shift", command=lambda:self.cover_shift())
        self.cover_shift_button.place(x=70, y=550)
        # Separator between buttons and other UI elements
        self.sep2 = ttk.Separator(self.root, orient='horizontal')
        self.sep2.place(x=0, y=600, relwidth=1.2)
        # Button for resetting data in selected row
        self.reset_row_button = CTkButton(self.root, text="Reset Data in Selected Row",width=16, command=lambda: self.reset_selected_row())
        self.reset_row_button.place(x=180, y=500)
        # Label for indicating the need to save data
        self.save_label = tk.Label(self.root, text="You must save it to the database before you can export it")
        self.save_label.place(x=50, y=610)
        # Button for saving data to the database
        self.save_to_db_button = CTkButton(self.root, text="Save to Database",width=16, command=lambda: self.save_to_database())
        self.save_to_db_button.place(x=30, y=640)
        # Counter for CSV export
        self.csv_count = 0
        # Temporary root for treeview
        temp_root = Toplevel(self.root)
        # Create treeview
        self.create_treeview(temp_root)

    # Function to populate the dropdown menu with values from the linked list
    def populate_combobox(self, linked_list):
        current = linked_list.head
        values = []
        while current:
            values.append(current.data)
            current = current.next
        #adds the values to the combobox
        self.new_person_combobox['values'] = values
        # Check if there are any values in the Combobox
        if values:
            # Set the value of the dropdown menu to the data of the first node
            self.new_person_combobox.set(values[0])

    # Function to clear and repopulate the dropdown menu with new values
    def refresh_combobox(self, linked_list):
        #if the linked list is empty
        if linked_list.head is None:
            #clears the current values
            self.new_person_combobox['values'] = ()
            self.new_person_combobox.set("")
        else:
            #populate the dropdown menu with new values
            self.populate_combobox(linked_list)
            first_node_data = linked_list.head.data
            self.new_person_combobox.set(first_node_data)

    def cover_shift(self):
        # Get the selected staff name and day
        staff_name = self.staff_name_combobox.get()
        day = self.day_combobox.get()
        # Store the list of children in the Treeview
        items = self.my_tree.get_children()
        # Check for "Unavailable" in the selected staff's day column
        for index, item in enumerate(items):
            item_values = self.my_tree.item(item)["values"]
            if item_values and item_values[0] == staff_name:
                # Determine the index of the selected day in the values list
                day_index = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"].index(day) + 1
                # Check if the selected staff is already covering in this store
                if item_values[0] in self.other_store:
                    tk.messagebox.showerror("Error", "They are already covering in this store")
                    return
                # Check if the selected staff is marked as "Unavailable" on this day
                if "Unavailable" in item_values[day_index]:
                    tk.messagebox.showerror("Error", "They are unavailable on this day.")
                    return
                else:
                    # Fetch secondary store information
                    primid = self.store_number[0][0]
                    conn = sqlite3.connect("Rota_Maker.db")
                    cursor = conn.cursor()
                    try:
                        secid = cursor.execute("SELECT Closest_Store FROM Account WHERE Primary_Store = (?)", (primid,)).fetchall()[0][0]
                        storename = cursor.execute("SELECT Store_Name FROM Stores WHERE StoreID = (?)", (secid,)).fetchall()[0][0]
                    except:
                        tk.messagebox.showerror("Error", "There is no one working at this store")
                        return
                    # Extract abbreviated store name
                    storename = storename.split()
                    if len(storename) == 2:
                        abbstore = storename[0][0].upper() + storename[1][0].upper()
                    elif len(storename) == 1:
                        abbstore = storename[0][:2].upper()
                    # Get the next item in the Treeview
                    next_item = items[index + 1] if index + 1 < len(items) else None
                    # Update the value in the Treeview to the abbreviated store name
                    self.my_tree.set(next_item, column=day_index, value=abbstore)
                    # Extract total_hours before resetting the cell value
                    total_hours = float(item_values[day_index].split()[-1])
                    # Set the value in the current row and column to "0.00"
                    self.my_tree.set(item, day_index, value="0.00")
                    last_column_index = len(item_values) - 1
                    # Decrement the values in the second last row but the same column of the treeview
                    budget_row_item = self.my_tree.get_children()[-2]
                    last_column_value = float(self.my_tree.item(budget_row_item)["values"][last_column_index])
                    self.my_tree.set(budget_row_item, last_column_index, value=f"{last_column_value - total_hours:.2f}")
                    budget_hrs_value = float(self.my_tree.item(budget_row_item)["values"][day_index])
                    self.my_tree.set(budget_row_item, day_index, value=f"{budget_hrs_value - total_hours:.2f}")
                    # Decrement the values in the last column but the same row of the treeview
                    last_column_value = float(self.my_tree.item(item)["values"][last_column_index])
                    self.my_tree.set(item, last_column_index, value=f"{last_column_value - total_hours:.2f}")

    def remove_selected(self):
        # Get the currently selected items in the Treeview
        selected_items = self.my_tree.selection()
        if selected_items:
            # Get the first selected item
            selected_item = selected_items[0]
            # Get the values of the selected item
            selected_values = self.my_tree.item(selected_item, "value")
            # Check if the selected item is an "Actual" or "Sign" entry
            if selected_values[0] in ("Actual", "Sign"):
                tk.messagebox.showerror("Error", "You must click on their name to remove them")
                return
            # Check if the selected staff member is from another store
            if not selected_values[0] in self.other_store:
                tk.messagebox.showerror("Error", "You can only remove a staff member of a different store")
                return
            name = selected_values[0]
            # Remove the staff member from the list of first names
            self.first_names.remove(name)
            # Update the values in the staff name combobox
            self.staff_name_combobox.configure(values=self.first_names)
            # Add the removed staff member back to the linked list
            self.staff_names_linked_list.append(name)
            # Refresh the staff name combobox
            self.refresh_combobox(self.staff_names_linked_list)
            # Find the index of the selected row to remove
            index_to_remove = None
            counter = 0
            for item in self.my_tree.get_children():
                values = self.my_tree.item(item, "values")
                if values == selected_values:
                    index_to_remove = counter
                    break
                counter += 1
            # Remove the selected row and the two rows below it
            if index_to_remove is not None:
                index_to_remove += 1  # Adjust index to start from 1 instead of 0
                items_to_remove = [selected_item]
                for _ in range(2):
                    next_item = self.my_tree.next(selected_item)
                    if next_item:
                        items_to_remove.append(next_item)
                        selected_item = next_item
                for item in items_to_remove:
                    self.my_tree.delete(item)
                # Update index numbers of rows below the removed rows
                items = self.my_tree.get_children()
                for item in items:
                    index_text = self.my_tree.item(item, "text")
                    if index_text:
                        index = int(index_text)
                        if index > index_to_remove:
                            new_index = index - 3
                            self.my_tree.item(item, text=str(new_index))
            # Update height and geometry of the window
            self.h = self.h - 3
            self.height = (self.h)*22
            self.my_tree.configure(height=self.h)
            self.root1.geometry(f"1380x{self.height}")
        else:
            tk.messagebox.showerror("Error", "There is no row selected")

    def add_new_person(self, new_person_name):
        # Iterate through each row in the Treeview
        for child in self.my_tree.get_children():
            values = self.my_tree.item(child, "values")
            # Check if the new person's name is blank
            if not new_person_name:
                tk.messagebox.showerror("Error", f"You cannot add a blank name")
                return
            # Check if the new person's name already exists in the rota
            if values[0] == new_person_name:
                tk.messagebox.showerror("Error", f"{new_person_name} is already in this rota")
                return
        # Append the new person's name to the list of first names
        self.first_names.append(new_person_name)
        # Update the values in the staff name combobox
        self.staff_name_combobox.configure(values=self.first_names)
        # Remove the new person's name from the linked list
        self.staff_names_linked_list.remove(new_person_name)
        try:
            # Refresh the staff name combobox if the linked list is not empty
            if self.staff_names_linked_list.head:
                self.refresh_combobox(self.staff_names_linked_list)
        except:
            pass
        else:
            self.refresh_combobox(self.staff_names_linked_list)
        # Get the index of the "Budget Hrs" row
        count = None
        for index, value in enumerate(self.my_tree.get_children()):
            if self.my_tree.item(value, "values")[0] == "Budget Hrs":
                count = index - 1
                break
        num = self.store_number1[0]
        # Query database to retrieve working days for the new person
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        person_name = new_person_name.lower()
        working_days = cursor.execute("""SELECT Monday_Working, Tuesday_Working, Wednesday_Working, Thursday_Working, Friday_Working,
                                      Saturday_Working, Sunday_Working FROM Account WHERE UPPER(SUBSTR(Name, 1, INSTR(Name, ' ') - 1)) LIKE UPPER(?) AND Closest_Store = ?""", (person_name, num)).fetchall()
        conn.close()
        val = [new_person_name]
        # Populate the values list for the new person with their working days
        for i in range(len(working_days[0])):
            if 0 == working_days[0][i]:
                val.append("Unavailable")
            else:
                val.append("0.00")
        val.append("0.00")
        if count is not None:
            # Insert the new person just before the "Budget Hrs" row
            self.my_tree.insert("", count, values=val, tags=('bold',))
            val1 = ["Actual", "", "", "", "", "", "", "", ""]
            val2 = ["Sign", "", "", "", "", "", "", "", ""]
            self.my_tree.insert("", count+1, text="", values=val1, tags=("font_size",))
            self.my_tree.insert("", count+2, text="", values=val2, tags=("font_size",))
        else:
            # "Budget Hrs" row not found, insert at the end
            self.my_tree.insert("", "end", values=(new_person_name, "", "", "", "", "", "", "", ""), tags=('bold',))
            val1 = ["Actual", "", "", "", "", "", "", "", ""]
            val2 = ["Sign", "", "", "", "", "", "", "", ""]
            self.my_tree.insert("", "end", text="", values=val1, tags=("font_size",))
            self.my_tree.insert("", "end", text="", values=val2, tags=("font_size",))
        # Align columns to the right
        for col in range(1, 9):
            self.my_tree.column(col, anchor="e")
        # Update height and geometry of the window
        self.h = self.h + 3
        self.height = (self.h)*22
        self.my_tree.configure(height=self.h)
        self.root1.geometry(f"1380x{self.height}")

    def undo_date_clicked(self):
        # Reset GUI elements to their initial state
        self.date_dropdown.configure(state="readonly")
        self.save_date_button.place(x=205, y=85)
        self.undo_date_button.place_forget()
        self.staff_name_label.place_forget()
        self.day_label.place_forget()
        self.start_time_label.place_forget()
        self.end_time_label.place_forget()
        self.new_person_label.place_forget()
        self.staff_name_combobox.place_forget()
        self.day_combobox.place_forget()
        self.start_time_combobox.place_forget()
        self.end_time_combobox.place_forget()
        self.new_person_combobox.place_forget()
        self.remove_button.place_forget()
        self.add_new_person_button.place_forget()
        self.add_button.place_forget()
        self.reset_row_button.place_forget()
        self.save_label.place_forget()
        self.save_to_db_button.place_forget()
        self.sep1.place_forget()
        self.sep2.place_forget()
        self.cover_shift_button.place_forget()
        if self.csv_count == 1:
            self.export_excel_button.place_forget()
            self.convert_button_png.place_forget()
        self.root1.destroy()

    def create_treeview(self, root):
        # Initialize the main window for the Treeview
        self.root1 = root
        # Function to handle window closing
        def upon_closing():
            # Prompt user for confirmation before closing
            answer = tk.messagebox.askyesno(title="Confirmation", message="Are you sure you want to close the spreadsheet")
            if not answer:
                return
            else:
                self.undo_date_clicked()
        self.root1.protocol("WM_DELETE_WINDOW", upon_closing)  # Bind closing event
        temp_number = str(self.store_number[0][0])
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        temp_names = cursor.execute("SELECT Name FROM Account WHERE Primary_Store = (?)", (temp_number)).fetchall()
        conn.close()
        temp_names_len = len(temp_names)
        self.h = ((temp_names_len+1) *3 )+1
        self.height = (self.h*22)
        if self.height > 900:
            self.height = 900
        self.root1.geometry(f"1380x{self.height}+500+10")
        self.root1.title("Rota Window")
        self.root1.resizable(0,0)
        self.root1.configure(bg="#272727")
        # Create the Treeview widget
        self.my_tree = ttk.Treeview(self.root1, height=self.h)
        self.my_tree["columns"] = ("Staff Name", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday", "Total Hrs")
        # Configure columns
        self.my_tree.column("#0", width=0, stretch=False)
        self.my_tree.column("Staff Name", minwidth=178,  width=178, anchor=CENTER, stretch=False)
        self.my_tree.column("Monday", minwidth=156,   width=156, stretch=False)
        self.my_tree.column("Tuesday",  minwidth=156, width=156, stretch=False)
        self.my_tree.column("Wednesday", minwidth=156,  width=156, stretch=False)
        self.my_tree.column("Thursday", minwidth=156,  width=156, stretch=False)
        self.my_tree.column("Friday", minwidth=156,   width=156, stretch=False)
        self.my_tree.column("Saturday",minwidth=156,  width=156, stretch=False)
        self.my_tree.column("Sunday", minwidth=156, width=156, stretch=False)
        self.my_tree.column("Total Hrs", minwidth=90,  width=90, stretch=False)
        # Configure column headings
        self.my_tree.heading("Staff Name", text="Staff Name", anchor=CENTER)
        self.my_tree.heading("Monday", text="Monday", anchor=CENTER)
        self.my_tree.heading("Tuesday", text="Tuesday", anchor=CENTER)
        self.my_tree.heading("Wednesday", text="Wednesday", anchor=CENTER)
        self.my_tree.heading("Thursday", text="Thursday", anchor=CENTER)
        self.my_tree.heading("Friday", text="Friday", anchor=CENTER)
        self.my_tree.heading("Saturday",text="Saturday", anchor=CENTER)
        self.my_tree.heading("Sunday", text="Sunday", anchor=CENTER)
        self.my_tree.heading("Total Hrs", text="Total Hrs", anchor=CENTER)
        # Apply tags to columns
        self.my_tree.tag_configure('bold', font=('Arial', 10, 'bold'))
        self.my_tree.tag_configure('font_size', font=('Arial', 9))
        # Determine table name for database query
        date = str(self.date)
        store = self.store.replace(" ", "_")
        date = self.date.replace("/", "_")
        self.table_name = f"unfinished_rota_{store}_{date}"
        try:
            conn = sqlite3.connect('Rota_Maker.db')
            cursor = conn.cursor()
            cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name=?", (self.table_name,))
            existing_table = cursor.fetchone()
            conn.close()
            if existing_table:
                # Prompt user to load existing rota data
                answer = tk.messagebox.askquestion("Load Data", """A Rota for this specific store and date already exists.
Do you want to load the rota from the database?""")
                if answer == 'yes':
                    # Load data from the existing table into the Treeview
                    self.load_data_from_database()
                else:
                    # Load data as normal
                    self.load_data_as_normal()
            else:
                # Table doesn't exist, load data as normal
                self.load_data_as_normal()
        except Exception as e:
            # Show error message if an error occurs
            tk.messagebox.showerror("Error", f"An error occurred: {str(e)}")
        # Pack the Treeview widget onto the new window
        self.my_tree.pack()

    def load_data_from_database(self):
        # Connect to the database
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        try:
            # Attempt to retrieve store number
            self.store_number1 = self.store_number[0]
        except:
            pass
        else:
            # Retrieve names associated with the store
            names_temp = cursor.execute("SELECT Name FROM Account WHERE Primary_Store = (?)", (self.store_number1)).fetchall()
            self.names = []
            array = []
            # Iterate over retrieved names
            for i in names_temp:
                temp = i[0]
                self.names.append(temp)
        # Load data from the existing table into the Treeview
        select_data_query = f"SELECT Staff_Name, Monday, Tuesday, Wednesday, Thursday, Friday, Saturday, Sunday, Total_Hrs FROM {self.table_name}"
        data_from_db = cursor.execute(select_data_query).fetchall()
        # Iterate over retrieved data and insert into Treeview
        for index, row in enumerate(data_from_db):
            # Apply 'bold' tag to staff name row and last row (budget hours)
            tags = ("bold",) if index % 3 == 0 or index == len(data_from_db) - 2 else ("font_size",)
            # Insert data into the Treeview
            self.my_tree.insert("", "end", values=row, tags=tags)
            # Anchor values to the right
            for col in range(1, 9):
                self.my_tree.column(col, anchor="e")
        # Close the database connection
        conn.close()

    def load_data_as_normal(self):
        try:
            # Attempt to retrieve store number
            self.store_number1 = self.store_number[0]
        except:
            pass
        else:
            # Connect to the database
            conn = sqlite3.connect('Rota_Maker.db')
            cursor = conn.cursor()
            # Retrieve names associated with the store
            names_temp = cursor.execute("SELECT Name FROM Account WHERE Primary_Store = (?)", (self.store_number1)).fetchall()
            self.names = []
            array = []
            # Iterate over retrieved names
            for i in names_temp:
                temp = i[0]
                self.names.append(temp)
            first_names = []
            # Iterate over names to extract first names
            for i in self.names:
                temp = []
                first_name=i.split()[0].capitalize()
                first_names.append(first_name)
                # Retrieve working days for each staff member
                working_days = cursor.execute("""SELECT Monday_Working, Tuesday_Working, Wednesday_Working, Thursday_Working, Friday_Working,
                                              Saturday_Working, Sunday_Working FROM Account WHERE Name = (?)""", (i,)).fetchall()
                val = working_days[0]
                val = str(val).split(",")
                temp.append(first_name)
                temp.append(val)
                array.append(temp)
            # Close the database connection
            conn.close()
            count = 0
            # Iterate over the array to insert data into Treeview
            for i in array:
                val = []
                val.append(i[0])
                for j in range(len(i[1])):
                    if "0" in i[1][j]:
                        val.append("Unavailable")
                    else:
                        val.append("0.00")
                val.append("0.00")
                val1 = ["Actual", "", "", "", "", "", "", "", ""]
                val2 = ["Sign", "", "", "", "", "", "", "", ""]
                self.my_tree.insert("", count, text="", values=val, tags=("bold",))
                self.my_tree.insert("", count+1, text="", values=val1, tags=("font_size",))
                self.my_tree.insert("", count+2, text="", values=val2, tags=("font_size",))
                count+= 3
        # Insert a space between staff and budget hours
        for i in range(1):
            self.my_tree.insert("", "end", text="", values=[""] * 9)
        # Insert budget hours and actual hours rows
        self.budget_hrs_values = ["Budget Hrs"] + ["0.00"] * 7 + ["0.00"]
        self.my_tree.insert("", "end", text="", values=self.budget_hrs_values, tags=("bold",))
        self.actual_hrs_values = ["Actual Hrs"] + [""] * 7 + [""]
        self.my_tree.insert("", "end", text="", values=self.actual_hrs_values)
        # Anchor values to the right
        for col in range(1, 9):
            self.my_tree.column(col, anchor="e")

    def add_data(self, staff_name, day, start_time, end_time):
        index = 0
        #Get each row in the treeview
        for item in self.my_tree.get_children():
            #get the values in that row
            item_values = self.my_tree.item(item)["values"]
            #iterate through until you get the row of the selected staff member
            if item_values and item_values[0] == staff_name:
                #get the specific day they wanted to add data to
                day_index = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"].index(day) + 1
                #check if they are unavailable on that day
                if "Unavailable" in item_values[day_index]:
                    #display an error message if they are unavailable
                    tk.messagebox.showerror("Error", "They are unavailable on this day.")
                    return
                ind = index + 1
                #get the values in the "Actual" row (below their times)
                vals = self.my_tree.item(self.my_tree.get_children()[ind])["values"]
                for i in range(1, 8):
                    if i == day_index:
                        #check if they are covering a shift in a different store
                        if vals[i] != "":
                            tk.messagebox.showerror("Error", f"They are covering a shift in {self.storename}")
                            return
                break
            index += 1
        try:
            #get the start and end times
            start_time_obj = datetime.datetime.strptime(start_time, "%H:%M")
            end_time_obj = datetime.datetime.strptime(end_time, "%H:%M")
            #check if start time is greater or equal to end time
            if end_time_obj <= start_time_obj:
                #display an error message
                tk.messagebox.showerror("Error", "End time must be greater than start time.")
                return
        except ValueError:
            #check if they have entered valid times and display an error message if not
            tk.messagebox.showerror("Error", "Invalid time format. Please use HH:MM.")
            return
        #get the time difference between start and end times
        time_difference = end_time_obj - start_time_obj
        #get the number of hours difference from that
        total_hours = time_difference.total_seconds() / 3600
        # Iterate through all rows in the treeview
        for item in self.my_tree.get_children():
            #get the values of each row
            item_values = self.my_tree.item(item)["values"]
            #get the selected staff's row
            if item_values and item_values[0] == staff_name:
                #get the seleceted day
                day_index = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"].index(day) + 1
                #add the selected start time, end time and total hours into the chosen cell in the treeview
                item_values[day_index] = f"{start_time}  ->  {end_time}    {total_hours:.2f}"
                #get the total hours for the staff member
                total_hrs = sum(float(item_values[i].split(" ")[-1]) if len(item_values[i].split(" ")) >= 2 else 0
                    for i in range(1, 8))
                #update the total hours column
                item_values[-1] = f"{total_hrs:.2f}"
                #update the values in the treeview
                self.my_tree.item(item, values=item_values)
        # Update the budget hours row
        budget_hrs_values = ["Budget Hrs"] + ["0.00"] * 7 + ["0.00"]
        for i in range(1, 8):
            total_hrs_for_day = 0
            #get each row in treeview
            for item in self.my_tree.get_children():
                #get each row's values
                item_values = self.my_tree.item(item)["values"]
                #check if there are values in that cell
                if item_values[i] and len(item_values[i].split(" ")) >= 2:
                    #add that to the total hours for the day
                    total_hrs_for_day += float(item_values[i].split(" ")[-1])
                else:
                    #don't add anything
                    total_hrs_for_day += 0
            #update the budget hrs row or leave it as "0.00"
            budget_hrs_values[i] = f"{total_hrs_for_day:.2f}" if total_hrs_for_day > 0 else "0.00"
        total_hrs_for_week = 0
        #iterate through each row in the treeview
        for item in self.my_tree.get_children():
            #get the values for that row
            item_values = self.my_tree.item(item)["values"]
            #remove any whitespace on the end of that value
            item_values[-1].strip()
            #look at every row except the "Budget Hrs" row
            if not item_values[0] == "Budget Hrs":
                #get the final cell on each row (the "Total Hrs" column)
                if item_values[-1]:
                    #add up the value in that cell with the pre-existing variable total_hrs_for_week
                    total_hrs_for_week = total_hrs_for_week + float(item_values[-1])
        #update the value in the "Total Hrs" column cell in the "Budget Hrs" row
        budget_hrs_values[-1] = f"{total_hrs_for_week:.2f}"
        #update the data in the treeview
        self.my_tree.item(self.my_tree.get_children()[-2], values=budget_hrs_values)
        for item in self.my_tree.get_children():
            item_values = self.my_tree.item(item)["values"]
        
    def reset_selected_row(self):
        # Check if a row is selected
        selected_item = self.my_tree.selection()
        if not selected_item:
            # Show a warning message if no row is selected
            tk.messagebox.showwarning("Warning", "Please select a row.")
            return
        # Get the values of the selected row as a list
        values = list(self.my_tree.item(selected_item, 'values'))
        # Reset the values in the days column to "0.00" or empty string
        if values[0] == "Actual":
            # Reset actual hours row
            for i in range(1, 8):
                values[i] = ""
            # Update the Treeview with the modified values
            self.my_tree.item(selected_item, values=tuple(values))
        elif values[0] not in ("Budget Hrs", "Actual Hrs", "Sign", ""):
            # Reset staff row
            for i in range(1, 8):  # Assuming columns 1 to 7 correspond to Monday to Sunday
                if not values[i] == "Unavailable":
                    values[i] = "0.00"
            # Update the Treeview with the modified values
            self.my_tree.item(selected_item, values=tuple(values))
            # Update the 'Total Hrs' value
            total_hrs = 0
            for value in values[1:8]:
                if not value == "Unavailable":
                    total_hrs += float(value)
            values[-1] = f"{total_hrs:.2f}"  # Format total hours to two decimal places
            # Update the Treeview with the modified 'Total Hrs' value
            self.my_tree.item(selected_item, values=tuple(values))
            # Update the 'Budget Hrs' value for the last row
            last_row_item = self.my_tree.get_children()[-2]
            budget_hrs_values = list(self.my_tree.item(last_row_item, 'values'))
            for i in range(1, 9):
                # Calculate the total hours for each day
                total_hrs_for_day = 0
                for item in self.my_tree.get_children():
                    item_values = self.my_tree.item(item)["values"]
                    if item_values[i] and len(item_values[i].split(" ")) >= 2:
                        total_hrs_for_day += float(item_values[i].split(" ")[-1])
                    else:
                        total_hrs_for_day += 0
                budget_hrs_values[i] = f"{total_hrs_for_day:.2f}" if total_hrs_for_day > 0 else "0.00"
            # Update the Treeview with the modified 'Budget Hrs' values
            self.my_tree.item(last_row_item, values=tuple(budget_hrs_values))
        else:
            # Show an error message if the row cannot be deleted
            tk.messagebox.showerror(title="Error", message="You cannot delete this row of data")

    def save_to_database(self):
        # Prepare table name based on date and store
        date = str(self.date)
        store = self.store.replace(" ", "_")
        date = self.date.replace("/", "_")
        if self.admin == 4:
            self.table_name = f"finished_rota_{store}_{date}"
        else:
            self.table_name = f"unfinished_rota_{store}_{date}"
        # Confirm saving with user
        answer = tk.messagebox.askyesno(title="Confirmation", message="Are you sure you want to save the rota in the current state?")
        if not answer:
            return
        else:
            try:
                # Open a connection to the SQLite database
                conn = sqlite3.connect('Rota_Maker.db')
                cursor = conn.cursor()
                # Check if the table already exists
                cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name=?", (self.table_name,))
                existing_table = cursor.fetchone()
                if existing_table:
                    # Confirm override if table exists
                    answer = tk.messagebox.askyesno(title="Confirmation", message="This will override the previous save. Are you sure you want to continue?")
                    if not answer:
                        return
                    else:
                        # Delete existing data and reset id column if confirmed
                        cursor.execute(f"DELETE FROM {self.table_name}")
                        cursor.execute(f"DELETE FROM sqlite_sequence WHERE name=?", (self.table_name,))
                else:
                    # Create the table if it doesn't exist
                    create_table_query = f'''
                    CREATE TABLE IF NOT EXISTS {self.table_name} (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        Staff_Name TEXT,
                        Monday TEXT,
                        Tuesday TEXT,
                        Wednesday TEXT,
                        Thursday TEXT,
                        Friday TEXT,
                        Saturday TEXT,
                        Sunday TEXT,
                        Total_Hrs TEXT
                    );
                    '''
                    cursor.execute(create_table_query)
                # Insert data into the table
                for item in self.my_tree.get_children():
                    item_values = self.my_tree.item(item)["values"]
                    if item_values:
                        cursor.execute(f'''
                        INSERT INTO {self.table_name} (Staff_Name, Monday, Tuesday, Wednesday, Thursday, Friday, Saturday, Sunday, Total_Hrs)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', item_values)
                # Commit changes and close the connection
                conn.commit()
                conn.execute("VACUUM")  # Reset the auto-increment counter
                tk.messagebox.showinfo("Success", "Data saved to the database.")
                conn.close()
                # Enable export buttons
                self.csv_count = 1
                self.export_excel_button = CTkButton(self.root, text="Export to Excel",width=16, command=lambda: self.export_to_excel())
                self.export_excel_button.place(x=170, y=640)
                self.convert_button_png = CTkButton(self.root, text="Convert to PNG",width=16, command=lambda: self.convert_treeview_to_png())
                self.convert_button_png.place(x=280, y=640)
            except Exception as e:
                # Show error message if an error occurs
                tk.messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def export_to_excel(self):
        # Prepare table name based on date and store
        date = str(self.date)
        store = self.store.replace(" ", "_")
        date = self.date.replace("/", "_")
        table_name = f"unfinished_rota_{store}_{date}"
        
        try:
            # Connect to the SQLite database
            conn = sqlite3.connect('Rota_Maker.db')
            cursor = conn.cursor()
            
            # Fetch column information and data from the table
            cursor.execute(f"PRAGMA table_info('{table_name}')")
            column_info = cursor.fetchall()
            columns_without_id = [column[1] for column in column_info if column[1] != 'id']
            cursor.execute(f"SELECT {', '.join(columns_without_id)} FROM {table_name}")
            rows = cursor.fetchall()
            conn.close()
            
            # Get the file path for saving the Excel file
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
            if not file_path:
                return  
            
            # Write data to the Excel file
            workbook = Workbook()
            sheet = workbook.active
            
            # Write header without the 'id' column
            for col, header in enumerate(columns_without_id, start=1):
                sheet.cell(row=1, column=col, value=header)
            
            # Write data rows
            for row_idx, row in enumerate(rows, start=2):
                for col_idx, value in enumerate(row, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Apply bold formatting to specific rows
            bold_rows = [(i * 3) + 2 for i in range(len(self.names))]
            bold_rows.append((len(self.names) * 3) + 3)
            for row_idx in bold_rows:
                for col_idx in range(1, len(columns_without_id) + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.font = Font(bold=True)
            
            # Save the Excel file
            workbook.save(file_path)
            tk.messagebox.showinfo("Success", f"Data exported to {file_path}")
        except Exception as e:
            # Show error message if an error occurs
            tk.messagebox.showerror("Error", f"An error occurred: {str(e)}")
        finally:
            pass
    
    def convert_treeview_to_png(self):
        # Take a screenshot of the entire window
        x = self.my_tree.winfo_rootx()
        y = self.my_tree.winfo_rooty()
        x1 = x + self.my_tree.winfo_width()
        y1 = y + self.my_tree.winfo_height()
        image = ImageGrab.grab(bbox=(x, y, x1, y1))
        # Save the image as PNG
        file_path = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG Files", "*.png")])
        if not file_path:
            return  # User canceled the save dialog
        image.save(file_path)
        tk.messagebox.showinfo("Success", f"Treeview converted to PNG and saved to {file_path}")

class Next_Week:
    def __init__(self, root):
        # Initialize the class with the provided root window
        self.root = root
        # Variable to store the maximum height of the Treeview
        self.max_height = 0
        # Get the current date and time
        today = datetime.datetime.now()
        # Get the weekday (0 for Monday, 1 for Tuesday, ...)
        today_weekday = today.weekday()
        # Calculate the number of days until the next Monday
        days_until_next_monday = (7 - today_weekday) % 7
        # Check if today is after Thursday in the previous week and before Thursday in this week
        if today_weekday >= 3 and today_weekday <= 6:
            next_monday_date = today + timedelta(days=days_until_next_monday)
            self.mondaydate = next_monday_date.strftime("%d_%m_%Y")
        else:
            # If today is not after Thursday in the previous week or before Thursday in this week,
            # we find the next Monday after this week
            next_monday_date = today + timedelta(days=(7 - today_weekday))
            self.mondaydate = next_monday_date.strftime("%d_%m_%Y")
        # Convert the date format for displaying in the window title
        date = self.mondaydate.replace("_", "/")
        # Set the window title
        self.root.title(f"Next week's rotas, week beginning: {date}")
        # Set the window size and position
        self.root.geometry("1450x800+100+100")
        # Disable window resizing
        self.root.resizable(0, 0)
        # Set the background color of the window
        self.root.configure(bg="#272727")
        # Bind the closing event to the on_closing method
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        # Create a scrollable frame
        self.scrollable_frame = tk.Frame(root)
        self.scrollable_frame.pack(fill="both", expand=True)
        self.canvas = tk.Canvas(self.scrollable_frame)
        # Create a vertical scrollbar for the scrollable frame
        scrollbar = ttk.Scrollbar(self.scrollable_frame, orient="vertical", command=self.canvas.yview)
        # Create an inner frame to contain the scrollable content
        self.scrollable_frame_inner = tk.Frame(self.canvas)
        # Update the scroll region when the inner frame is resized
        self.scrollable_frame_inner.bind(
            "<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")
            )
        )
        # Create a window in the canvas to display the inner frame
        self.canvas.create_window((0, 0), window=self.scrollable_frame_inner, anchor="nw")
        # Configure the canvas to scroll vertically
        self.canvas.configure(yscrollcommand=scrollbar.set)
        # Pack the canvas and scrollbar
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        # Connect to the database
        self.connection = sqlite3.connect('Rota_Maker.db')
        self.cursor = self.connection.cursor()
        # Display table contents
        self.display_tables()
        # Bind mouse wheel events for scrolling
        self.canvas.bind("<MouseWheel>", self.on_mousewheel)

    def on_closing(self):
        # Decrement window counter by 1
        global next_week
        next_week -= 1
        # Close the window
        self.root.destroy()

    def display_tables(self):
        # Query to fetch table names
        self.cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = self.cursor.fetchall()
        x = 0
        # Check through each table in the database
        for table in tables:
            table_name = table[0]
            # Check if it is a finished rota, is for next Monday, and is a valid table
            if table_name.startswith('finished_rota') and self.mondaydate in table_name:
                store_name = table_name.split('finished_rota_')[1].split(f"_{self.mondaydate}")[0]
                if "_" in store_name:
                    store_name = store_name.replace("_", ' ')
                # Extract store name
                self.display_table_contents(table_name, store_name)
                x += 1
        if x == 0:
            self.show_error_label("Error: Currently no finished rotas.")

    def display_table_contents(self, table_name, store_name):
        # Query to fetch column names from table (excluding 'id')
        self.cursor.execute(f"PRAGMA table_info({table_name})")
        column_names = [row[1] for row in self.cursor.fetchall() if row[1] != 'id']
        # Create a label to display the store name
        store_label = tk.Label(self.scrollable_frame_inner, text=store_name, font=("Arial", 16))
        store_label.pack()
        # Fetch data from table (excluding 'id')
        self.cursor.execute(f"SELECT {', '.join(column_names)} FROM {table_name}")
        data = self.cursor.fetchall()
        # Count the number of rows
        num_rows = len(data)
        # Calculate the height of the Treeview
        self.h = ((num_rows + 1))
        self.height = min(self.h * 23, 900)  # Update self.height
        # Update max_height if self.height is greater
        self.max_height = max(self.max_height + 200, self.height)
        self.root.geometry(f"1450x{self.max_height}")
        # Create a Treeview widget
        tree = ttk.Treeview(self.scrollable_frame_inner, columns=column_names, show='headings', height=self.h)
        tree.pack(padx=20)
        style = ttk.Style()
        style.theme_use("clam")
        style.map("Treeview", background=[("selected", "#b5b5b5")])
        # Define columns and headers
        for col, col_name in enumerate(column_names):
            tree.heading(col_name, text=col_name)
            tree.column(col_name, width=int(0.95 * 156))  # Adjust width as needed
        # Insert data into Treeview
        for index, row in enumerate(data):
            tags = ()  # No special tags by default
            # Apply 'bold' tag to every fourth row starting from the first staff name
            if index % 3 == 0:  # Skip first row (headers)
                tags = ('bold',)
            if row[-2]:
                tags += ('bold',)
            tree.insert('', 'end', values=row, tags=tags)
        # Disable selection
        tree.bind('<Button-1>', lambda event: 'break')
        tree.bind('<Button-3>', lambda event: 'break')
        tree.bind('<Button-2>', lambda event: 'break')
        # Disable editing
        tree.bind('<Double-1>', lambda event: 'break')
        # Disable resizing columns
        tree.bind('<Motion>', lambda event: 'break')
        for col in range(1, 9):
            tree.column(col, anchor="e")
        # Configure tag for bold style
        tree.tag_configure('bold', font=('Arial', 10, 'bold'))
        
    def show_error_label(self, message):
        # Create and display an error label
        error_label = tk.Label(self.scrollable_frame_inner, text=message, fg="red")
        error_label.pack()
        
    def on_mousewheel(self, event):
        # Scroll the canvas vertically in response to mouse wheel events
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

class Workers_Info:
    def __init__(self, root):
        # Initialize the class with the provided root window
        self.root = root
        # Set window title, size, and properties
        self.root.title("Worker Info")
        self.root.geometry("400x500+100+500")
        self.root.resizable(False, False)
        self.root.configure(bg="#272727")
        # Function to handle window closing
        def on_closing():
            global info_counter
            info_counter -= 1  # Decrement counter when the window is closed
            self.root.destroy()
        self.root.protocol("WM_DELETE_WINDOW", on_closing)  # Bind closing event
        # Create refresh button
        self.refresh_button = CTkButton(self.root, text="Refresh", font=("Arial", 11), width=70, command=self.refresh)
        self.refresh_button.place(x=250, y=20)
        # Create store selection label and dropdown
        self.store_label = CTkLabel(master=self.root, text="Store:", font=("Arial", 16))
        self.store_label.place(x=20, y=20)
        # Fetch store names from the database
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        cursor.execute("SELECT Store_Name FROM Stores")
        stores_data = cursor.fetchall()
        conn.close()
        # Extract store names from the fetched data
        self.stores = [store[0] for store in stores_data]
        self.store_combobox = CTkComboBox(self.root,state="readonly", values=self.stores)
        self.store_combobox.set(self.stores[0])
        self.store_combobox.place(x=80, y=20)
        # Create label to display worker details
        self.worker_details_box = Label(master=self.root, width=40, height=15, wraplength=500, font=("Arial", 12), bg="#272727", fg="white", anchor="nw", justify="left")
        self.worker_details_box.place(x=20, y=100)
        self.current_person_index = 0
        # Canvas for the worker dots
        self.worker_dot_canvas = Canvas(self.root, width=300, height=30, bg="#272727", highlightthickness=0)
        self.worker_dot_canvas.place(x=30, y=60)
        # Buttons for navigating through workers
        self.previous_button = Button(master=self.root, text="Previous", command=self.show_previous_person, font=("Arial", 12))
        self.previous_button.place(x=20, y=390)
        self.next_button = Button(master=self.root, text="Next", command=self.show_next_person, font=("Arial", 12))
        self.next_button.place(x=160, y=390)
        self.update_worker_details()

    def refresh(self):
        # Refresh worker details
        self.update_worker_details()

    def update_worker_details(self, *args):
        # Get the selected store name
        store_name = self.store_combobox.get()
        # Get the corresponding store number
        store_number = self.stores.index(store_name) + 1
        # Connect to the database
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        # Retrieve worker details from the database
        worker_details = cursor.execute(f"""SELECT Name, Email, Birthday, Monday_Working, Tuesday_Working, Wednesday_Working, Thursday_Working, Friday_Working,
                                        Saturday_Working, Sunday_Working, Contract_Hours, Primary_Store, Closest_Store
                                        FROM Account a, Stores  b WHERE a.Primary_Store = b.StoreID AND StoreID = {store_number}""").fetchall()
        # Close the database connection
        conn.close()
        if worker_details:
            # Update the worker details attribute
            self.worker_details = worker_details
            # Reset the current person index to 0
            self.current_person_index = 0
            # Display details of the first worker
            self.show_person_details()
            # Update the worker dots to reflect the current worker
            self.update_worker_dots()
        else:
            # Display an error message
            self.worker_details_box.config(text="No worker details found for the selected store.", fg="white")

    def show_person_details(self):
        # Display details of the current worker
        person = self.worker_details[self.current_person_index]
        name, email, birthday, monday_working, tuesday_working, wednesday_working, thursday_working, friday_working, saturday_working, sunday_working, contract_hours, primary_store_id, closest_store_id = person
        name = name.split()
        x = name[0].capitalize()
        y = name[-1].capitalize()
        name = x + " " + y
        # Convert working days to "Yes" or "Unavailable"
        monday_working = "Yes" if monday_working == 1 else "Unavailable"
        tuesday_working = "Yes" if tuesday_working == 1 else "Unavailable"
        wednesday_working = "Yes" if wednesday_working == 1 else "Unavailable"
        thursday_working = "Yes" if thursday_working == 1 else "Unavailable"
        friday_working = "Yes" if friday_working == 1 else "Unavailable"
        saturday_working = "Yes" if saturday_working == 1 else "Unavailable"
        sunday_working = "Yes" if sunday_working == 1 else "Unavailable"
        # Get store names based on store IDs
        primary_store_name = self.stores[primary_store_id - 1]
        closest_store_name = self.stores[closest_store_id - 1]
        details_text = f"""Name: {name}
Email: {email}
Birthday: {birthday}

Contract Hours: {contract_hours}
Primary Store: {primary_store_name}
Closest Store: {closest_store_name}

Monday Working: {monday_working}
Tuesday Working: {tuesday_working}
Wednesday Working: {wednesday_working}
Thursday Working: {thursday_working}
Friday Working: {friday_working}
Saturday Working: {saturday_working}
Sunday Working: {sunday_working}"""
        self.worker_details_box.config(text=details_text, fg="white")

    def show_next_person(self):
        # Display details of the next worker
        # Check if there are worker details available
        if self.worker_details:
            # Update the current person index to point to the next person in the list
            self.current_person_index = (self.current_person_index + 1) % len(self.worker_details)
            # Display details of the next person
            self.show_person_details()
            # Update the worker dots to reflect the change
            self.update_worker_dots()

    def show_previous_person(self):
        # Display details of the previous worker
        # Check if there are worker details available
        if self.worker_details:
            # Update the current person index to point to the previous person in the list
            self.current_person_index = (self.current_person_index - 1) % len(self.worker_details)
            # Display details of the previous person
            self.show_person_details()
            # Update the worker dots to reflect the change
            self.update_worker_dots()

    def update_worker_dots(self):
        # Update worker dots based on current worker index
        self.worker_dot_canvas.delete("all")
        num_workers = len(self.worker_details)
        # Calculate dot positions
        dot_radius = 5
        dot_spacing = 20
        total_width = (dot_radius * 2 + dot_spacing) * num_workers
        start_x = (300 - total_width) // 2
        # Draw dots
        for i in range(num_workers):
            color = "#00EEFF" if i == self.current_person_index else "white"
            self.worker_dot_canvas.create_oval(start_x + i * (dot_radius * 2 + dot_spacing),
                                               5, start_x + i * (dot_radius * 2 + dot_spacing) + dot_radius * 2,
                                               5 + dot_radius * 2, fill=color, outline=color)

# Define the Account_Settings class
class Account_Settings:
    # Initialization method that takes the root window and email as parameters
    def __init__(self, root, email):
        # Assign the root window and email to instance variables
        self.root = root
        self.email = email
        # Set the title, geometry, and disable resizing for the root window
        self.root.title("Account Settings")
        self.root.geometry("400x600+200+100")
        self.root.resizable(False, False)
        # Set the background color of the root window
        self.root.configure(bg="#272727")
        # Define a function to handle window closing
        def on_closing():
            # Decrement the global account_settings_counter when the window is closed
            global account_settings_counter
            account_settings_counter -= 1
            # Destroy the root window
            self.root.destroy()
        # Bind the closing event to the on_closing function
        self.root.protocol("WM_DELETE_WINDOW", on_closing)
        # Connect to the SQLite database
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        # Retrieve account data from the database based on the email
        account_data = cursor.execute("""SELECT Name, Email, Password, Birthday, Monday_Working, Tuesday_Working, Wednesday_Working, Thursday_Working, Friday_Working,
                                      Saturday_Working, Sunday_Working, Contract_Hours, Primary_Store, Closest_Store  FROM Account WHERE Email = (?)""", (self.email,)).fetchall()[0]
        # Unpack the account data into instance variables
        self.name, self.email, self.password, self.birthday, self.monday_working, self.tuesday_working, self.wednesday_working, self.thursday_working, self.friday_working, self.saturday_working, self.sunday_working, self.contract_hours, self.primary_store_id, self.closest_store_id = account_data

        # Close database connection
        conn.close()
        # Create a label for the name field
        self.name_label = CTkLabel(self.root, text="Name:", font=("Arial", 14))
        self.name_label.place(x=30, y=40)
        # Create an entry field for the name, insert the name, and make it read-only
        self.name_entry = CTkEntry(self.root, font=("Arial", 15), width=150)
        self.name_entry.insert(END, self.name)
        self.name_entry.place(x=100, y=40)
        self.name_entry.configure(state=("readonly"))
        # Create an "Edit" button for the name field
        self.edit_button1 = CTkButton(self.root, text="Edit", width=50, height=20, command=lambda: self.edit1_pressed())
        self.edit_button1.place(x=275, y=40)
        # Create a label for the email field
        self.email_label = CTkLabel(self.root, text="Email:", font=("Arial", 14))
        self.email_label.place(x=30, y=80)
        # Create an entry field for the email, insert the email, and make it read-only
        self.email_entry = CTkEntry(self.root, font=("Arial", 15), width=150)
        self.email_entry.insert(END, self.email)
        self.email_entry.place(x=100, y=80)
        self.email_entry.configure(state=("readonly"))
        # Create a label for the password field
        self.password_label = CTkLabel(self.root, text="Password:", font=("Arial", 14))
        self.password_label.place(x=30, y=120)
        # Create a button to change the password
        self.edit_button2 = CTkButton(self.root, text="Change Password", width=150, font=("Arial", 14), height=30, command=lambda: self.edit2_pressed())
        self.edit_button2.place(x=100, y=120)
        # Define lists for months, days, and years
        self.months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
        self.days = [f"{i:02d}" for i in range(1, 32)]
        self.years = list(range(1900, date.today().year))[::-1]
        # Create StringVar instances for selected day, month, and year
        self.selected_day = tk.StringVar()
        self.selected_month = tk.StringVar()
        self.selected_year = tk.StringVar()
        # Split the birthday date into day, month, and year
        self.birthday_date = self.birthday.split("/")
        day = self.birthday_date[0]
        month = int(self.birthday_date[1]) - 1
        month = self.months[month]
        year = self.birthday_date[2]
        # Create a label for the month dropdown
        self.month_label = ttk.Label(root, text="Month:")
        self.month_label.place(x=10, y=185)
        # Create a combobox for the month dropdown
        self.month_combobox = ttk.Combobox(root, state="readonly", textvariable=self.selected_month, width=10, values=self.months)
        self.month_combobox.place(x=10, y=205)
        self.month_combobox.set(month)
        self.month_combobox.bind("<<ComboboxSelected>>", self.update_days)
        self.month_combobox.config(state="disable")
        # Create a label for the birthday
        self.birthday = ttk.Label(root, text=("Birthday"), font=("Arial", 13))
        self.birthday.place(x=100, y=155)
        # Create a label for the day dropdown
        self.day_label = ttk.Label(root, text="Day:")
        self.day_label.place(x=100, y=185)
        # Create a combobox for the day dropdown
        self.day_combobox = ttk.Combobox(root, state="readonly", textvariable=self.selected_day, width=10, values=self.days)
        self.day_combobox.place(x=100, y=205)
        self.day_combobox.set(day)
        self.day_combobox.config(state="disabled")
        # Create a label for the year dropdown
        self.year_label = ttk.Label(root, text="Year: ")
        self.year_label.place(x=190, y=185)
        # Create a combobox for the year dropdown
        self.year_combobox = ttk.Combobox(root, state="readonly", textvariable=self.selected_year, width=10, values=self.years)
        self.year_combobox.place(x=190, y=205)
        self.year_combobox.set(year)
        self.year_combobox.config(state="disabled")
        # Create an "Edit" button for the birthday
        self.edit_button3 = CTkButton(self.root, text="Edit", width=50, height=20, command=lambda: self.edit3_pressed())
        self.edit_button3.place(x=285, y=205)
        # Create a label for the days working
        self.working_label = CTkLabel(self.root, text="Days Working: ", font=("Arial", 16))
        self.working_label.place(x=110, y=240)
        # Create a switch for Monday
        self.switch_var1 = IntVar(value=self.monday_working)
        self.monday_switch = CTkSwitch(master=root, text="Monday", command=self.switcher1, variable=self.switch_var1, onvalue=1, offvalue=0)
        self.monday_switch.place(x=10, y=270)
        self.monday_switch.configure(state=("disabled"))
        # Create a switch for Tuesday
        self.switch_var2 = IntVar(value=self.tuesday_working)
        self.tuesday_switch = CTkSwitch(master=root, text="Tuesday", command=self.switcher2, variable=self.switch_var2, onvalue=1, offvalue=0)
        self.tuesday_switch.place(x=110, y=270)
        self.tuesday_switch.configure(state=("disabled"))
        # Create a switch for Wednesday
        self.switch_var3 = IntVar(value=self.wednesday_working)
        self.wednesday_switch = CTkSwitch(master=root, text="Wednesday", command=self.switcher3, variable=self.switch_var3, onvalue=1, offvalue=0)
        self.wednesday_switch.place(x=210, y=270)
        self.wednesday_switch.configure(state=("disabled"))
        # Create a switch for Thursday
        self.switch_var4 = IntVar(value=self.thursday_working)
        self.thursday_switch = CTkSwitch(master=root, text="Thursday", command=self.switcher4, variable=self.switch_var4, onvalue=1, offvalue=0)
        self.thursday_switch.place(x=10, y=310)
        self.thursday_switch.configure(state=("disabled"))
        # Create a switch for Friday
        self.switch_var5 = IntVar(value=self.friday_working)
        self.friday_switch = CTkSwitch(master=root, text="Friday", command=self.switcher5, variable=self.switch_var5, onvalue=1, offvalue=0)
        self.friday_switch.place(x=110, y=310)
        self.friday_switch.configure(state=("disabled"))
        # Create a switch for Saturday
        self.switch_var6 = IntVar(value=self.saturday_working)
        self.saturday_switch = CTkSwitch(master=root, text="Saturday", command=self.switcher6, variable=self.switch_var6, onvalue=1, offvalue=0)
        self.saturday_switch.place(x=10, y=350)
        self.saturday_switch.configure(state=("disabled"))
        # Create a switch for Sunday
        self.switch_var7 = IntVar(value=self.sunday_working)
        self.sunday_switch = CTkSwitch(master=root, text="Sunday", command=self.switcher7, variable=self.switch_var7, onvalue=1, offvalue=0)
        self.sunday_switch.place(x=110, y=350)
        self.sunday_switch.configure(state=("disabled"))
        # Create an "Edit" button for the days working
        self.edit_button4 = CTkButton(self.root, text="Edit", width=50, height=20, command=lambda: self.edit4_pressed())
        self.edit_button4.place(x=210, y=350)
        # Create a label for the contract hours
        self.contract_hrs_label = CTkLabel(master=root, text="Your Contract Hours per Week: ", font=("Arial", 16))
        self.contract_hrs_label.place(x=80, y=390)
        # Define a list of contract hours values
        self.contract_hours_val = ["8", "16", "24", "32", "40"]
        # Create a combobox for the contract hours
        self.contract_hrs_combobox = ttk.Combobox(master=root, values=self.contract_hours_val)
        self.contract_hrs_combobox.place(x=110, y=430)
        self.contract_hrs_combobox.set(self.contract_hours)
        self.contract_hrs_combobox.configure(state=("disabled"))
        # Create an "Edit" button for the contract hours
        self.edit_button5 = CTkButton(self.root, text="Edit", width=50, height=20, command=lambda: self.edit5_pressed())
        self.edit_button5.place(x=260, y=430)
        # Create a label for the primary store
        self.primary_store_label = CTkLabel(master=root, text="Your Current Store: ", font=("Arial", 14))
        self.primary_store_label.place(x=90, y=460)
        # Define a list of primary store values
        self.primary_store_val = ["South Kensington", "Battersea", "Hampstead", "East Dulwich", "Chiswick", "Stoke Newington", "Wimbledon"]
        # Create a combobox for the primary store
        self.primary_store_combobox = ttk.Combobox(master=root, values=self.primary_store_val, font=("Helvetica", 9))
        self.primary_store_combobox.set(self.primary_store_val[self.primary_store_id-1])
        self.primary_store_combobox.configure(state=("disabled"))
        self.primary_store_combobox.place(x=90, y=490)
        # Create an "Edit" button for the primary store
        self.edit_button6 = CTkButton(self.root, text="Edit", width=50, height=20, command=lambda: self.edit6_pressed())
        self.edit_button6.place(x=280, y=490)

    def update_days(self, event):
        # Get the currently selected month from the month combobox
        selected_month = self.selected_month.get()
        # Set the default number of days in a month to 31
        days_in_month = 31
        # Check if the selected month is April, June, September, or November
        if selected_month in ["April", "June", "September", "November"]:
            # If so, set the number of days in the month to 30
            days_in_month = 30
        # Check if the selected month is February
        elif selected_month == "February":
            # Get the currently selected year from the year combobox
            selected_year = int(self.selected_year.get())
            # Check if the year is a leap year
            if (selected_year % 4 == 0 and selected_year % 100 != 0) or (selected_year % 400 == 0):
                # If it's a leap year, set the number of days in February to 29
                days_in_month = 29
            else:
                # If it's not a leap year, set the number of days in February to 28
                days_in_month = 28
        # Create a list of days based on the calculated number of days in the month
        self.days = list(range(1, days_in_month + 1))
        # Update the values in the day combobox with the new list of days
        self.day_combobox['values'] = self.days
        # Set the selected day in the day combobox to the first day of the month
        self.day_combobox.set(self.days[0])

    def edit1_pressed(self):
        # Set the state of the name entry to "normal" (editable)
        self.name_entry.configure(state="normal")
        # Create a "Save" button for the name entry
        self.save_button1 = CTkButton(self.root, text="Save", width=50, height=20, command=lambda: self.save_button1_pressed())
        self.save_button1.place(x=340, y=40)
        # Remove the "Edit" button for the name entry
        self.edit_button1.place_forget()
        # Create a "Cancel" button for the name entry
        self.cancel_button1 = CTkButton(self.root, text="Cancel", width=50, height=20, command=lambda: self.cancel_button1_pressed())
        self.cancel_button1.place(x=275, y=40)

    def cancel_button1_pressed(self):
        # Remove the "Cancel" and "Save" buttons for the name entry
        self.cancel_button1.place_forget()
        self.save_button1.place_forget()
        # Re-display the "Edit" button for the name entry
        self.edit_button1.place(x=275, y=40)
        # Connect to the SQLite database
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        # Retrieve the name from the database based on the email
        name = str(cursor.execute("SELECT Name FROM Account WHERE Email = (?)", (self.email,)).fetchall()[0][0])
        conn.close()
        # Clear the name entry field
        self.name_entry.delete(0, END)
        # Insert the name from the database into the name entry field
        self.name_entry.insert(END, name)
        # Set the state of the name entry to "readonly"
        self.name_entry.configure(state=("readonly"))

    def save_button1_pressed(self):
        # Get the entered name from the name entry field
        self.fullname = self.name_entry.get()
        # Check if the name is not empty and contains a space
        if self.fullname and " " in self.fullname:
            # Check if the name doesn't contain any digits
            if not re.search(r'\d', self.fullname):
                # Show a confirmation dialog
                answer = tk.messagebox.askyesno(title="Confirmation", message="Are you sure you would like update your name?")
                if answer:
                    # If the user confirms, connect to the database
                    conn = sqlite3.connect('Rota_Maker.db')
                    cursor = conn.cursor()
                    # Update the name in the database for the current email
                    cursor.execute("UPDATE Account SET Name = (?) WHERE Email = (?)", (self.fullname, self.email))
                    conn.commit()
                    conn.close()
                    # Show a success message
                    tk.messagebox.showinfo(title="Success", message="Name saved")
                    # Set the name entry to read-only, re-display the "Edit" button, and remove the "Save" and "Cancel" buttons
                    self.name_entry.configure(state=("readonly"))
                    self.edit_button1.place(x=275, y=40)
                    self.save_button1.place_forget()
                    self.cancel_button1.place_forget()
                else:
                    # If the user cancels, do nothing
                    return
            else:
                # If the name contains digits, show an error message
                tk.messagebox.showerror(title="Error", message="Invalid Name")
        else:
            # If the name is empty or doesn't contain a space, show an error message
            tk.messagebox.showerror(title="Error", message="Invalid Name")
    
    def edit2_pressed(self):
        # Show a confirmation dialog to update the password
        answer = tk.messagebox.askyesno(title="Confirmation", message="Are you sure you want to update your password?")
        if answer:
            # If the user confirms, create a new toplevel window
            self.root1 = CTkToplevel(self.root)
            self.root1.geometry("400x300")
            # Create an entry field to enter the current password
            self.current_password_entry = CTkEntry(master=self.root1, placeholder_text="Current Password", font=("Helvetica", 15), width=300, show="")
            self.current_password_entry.place(x=50, y=50)
            # Create a "Save" button to proceed with the password update
            self.save_button = CTkButton(self.root1, text="Save", font=("Arial", 16), width=90, corner_radius=40, command=lambda: self.save_clicked())
            self.save_button.place(x=155, y=100)
        else:
            # If the user cancels, do nothing
            return

    def save_clicked(self):
        # Get the current password entered by the user
        self.current_password = self.current_password_entry.get()
        if self.current_password:
            # Check if the entered current password matches the stored password
            if bcrypt.checkpw(self.current_password.encode("utf-8"), self.password):
                # If the current password is correct, make the current password entry read-only and remove the "Save" button
                self.current_password_entry.configure(state=("readonly"))
                self.save_button.place_forget()
                # Create entry fields for the new password and confirm new password
                self.password_entry = CTkEntry(master=self.root1, placeholder_text="New Password", font=("Helvetica", 15), width=300, show="")
                self.password_entry.place(x=50, y=140)
                self.confirm_password_entry = CTkEntry(master=self.root1, placeholder_text="Confirm New Password", font=("Helvetica", 15), width=300, show="")
                self.confirm_password_entry.place(x=50, y=190)
                # Create a "Save" button to save the new password
                self.save_password = CTkButton(self.root1, text="Save", corner_radius=40, font=("Arial", 16), width=90, command=lambda: self.save_password_clicked())
                self.save_password.place(x=155, y=240)
            else:
                # If the current password is incorrect, show an error message
                tk.messagebox.showerror(title="Error", message="Incorrect current password")
        else:
            # If the current password field is empty, show an error message
            tk.messagebox.showerror(title="Error", message="Current Password is empty")

    def save_password_clicked(self):
        # Get the new password and confirm new password entered by the user
        self.new_password = self.password_entry.get()
        self.confirm_password = self.confirm_password_entry.get()
        if self.new_password and self.confirm_password:
            if self.new_password == self.confirm_password:
                if self.new_password != self.current_password:
                    if len(self.new_password) >= 6:
                        # Hash the new password using bcrypt
                        self.new_hashed_password = bcrypt.hashpw((self.new_password.encode("utf-8")), bcrypt.gensalt())
                        # Update the password in the database
                        conn = sqlite3.connect('Rota_Maker.db')
                        cursor = conn.cursor()
                        cursor.execute("UPDATE Account SET Password = (?) WHERE Email = (?)", (self.new_hashed_password, self.email))
                        conn.commit()
                        conn.close()
                        # Show a success message and update the password attribute
                        tk.messagebox.showinfo(title="Success", message="Password changed")
                        self.password = self.new_hashed_password
                        # Close the toplevel window
                        self.root1.destroy()
                    else:
                        # If the new password is less than 6 characters, show an error message
                        tk.messagebox.showerror(title="Error", message="New Password must be 6 or more characters")
                else:
                    # If the new password is the same as the old password, show an error message
                    tk.messagebox.showerror(title="Error", message="New Password cannot be the same as old password")
            else:
                # If the new password and confirm new password don't match, show an error message
                tk.messagebox.showerror(title="Error", message="Passwords do not match")
        else:
            # If either the new password or confirm new password field is empty, show an error message
            tk.messagebox.showerror(title="Error", message="Fill out empty fields")

    def edit3_pressed(self):
        # Set the month, day, and year comboboxes to read-only
        self.month_combobox.config(state="readonly")
        self.day_combobox.config(state="readonly")
        self.year_combobox.config(state="readonly")
        # Remove the "Edit" button for the birthday
        self.edit_button3.place_forget()
        # Create a "Cancel" button for the birthday
        self.cancel_button3 = CTkButton(self.root, text="Cancel", width=50, height=20, command=lambda: self.cancel_button3_pressed())
        self.cancel_button3.place(x=285, y=195)
        # Create a "Save" button for the birthday
        self.save_button3 = CTkButton(self.root, text="Save", width=50, height=20, command=lambda: self.save_button3_pressed())
        self.save_button3.place(x=345, y=195)

    def cancel_button3_pressed(self):
        # Connect to the SQLite database
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        # Retrieve the birthday from the database based on the email
        birthday = str(cursor.execute("SELECT Birthday FROM Account WHERE Email = (?)", (self.email,)).fetchall()[0][0])
        conn.close()
        # Split the birthday into day, month, and year
        birthday_date = birthday.split("/")
        day = birthday_date[0]
        month = int(birthday_date[1]) - 1
        month = self.months[month]
        year = birthday_date[2]
        # Set the month, day, and year comboboxes with the retrieved values
        self.month_combobox.set(month)
        self.day_combobox.set(day)
        self.year_combobox.set(year)
        # Set the month, day, and year comboboxes to disabled
        self.month_combobox.config(state="disabled")
        self.day_combobox.config(state="disabled")
        self.year_combobox.config(state="disabled")
        # Re-display the "Edit" button for the birthday
        self.edit_button3.place(x=285, y=195)
        # Remove the "Cancel" and "Save" buttons for the birthday
        self.cancel_button3.place_forget()
        self.save_button3.place_forget()

    def save_button3_pressed(self):
        # Show a confirmation dialog to change the birthday
        answer = tk.messagebox.askyesno(title="Confirmation", message="Are you sure you want to change your birthday?")
        if answer:
            # Get the selected day, month, and year from the comboboxes
            self.day = str(self.day_combobox.get())
            self.month = str(self.month_combobox.get())
            self.year = str(self.year_combobox.get())
            # Create a dictionary to map month names to numbers
            self.months_dict = {"January": "01", "February": "02", "March": "03", "April": "04", "May": "05", "June": "06",
                                "July": "07", "August": "08", "September": "09", "October": "10", "November": "11", "December": "12"}
            # Convert the month name to its corresponding number
            self.month = self.months_dict[self.month]
            # Construct the new birthday string
            self.birthday = self.day+"/"+self.month+"/"+self.year
            # Connect to the SQLite database
            conn = sqlite3.connect('Rota_Maker.db')
            cursor = conn.cursor()
            # Update the birthday in the database for the current email
            cursor.execute("UPDATE Account SET Birthday = (?) WHERE Email = (?)", (self.birthday, self.email))
            conn.commit()
            conn.close()
            # Show a success message
            tk.messagebox.showinfo(title="Success", message="Birthday has been updated")
            # Call the cancel_button3_pressed method to reset the comboboxes and buttons
            self.cancel_button3_pressed()
        else:
            # If the user cancels, do nothing
            return

    def switcher1(self):
        pass

    def switcher2(self):
        pass

    def switcher3(self):
        pass

    def switcher4(self):
        pass

    def switcher5(self):
        pass
            
    def switcher6(self):
        pass
                    
    def switcher7(self):
        pass

    def edit4_pressed(self):
        # Set the state of all day switches to "normal" (editable)
        self.monday_switch.configure(state=("normal"))
        self.tuesday_switch.configure(state=("normal"))
        self.tuesday_switch.configure(state=("normal"))
        self.wednesday_switch.configure(state=("normal"))
        self.thursday_switch.configure(state=("normal"))
        self.friday_switch.configure(state=("normal"))
        self.saturday_switch.configure(state=("normal"))
        self.sunday_switch.configure(state=("normal"))
        # Remove the "Edit" button for the days working
        self.edit_button4.place_forget()
        # Create a "Cancel" button for the days working
        self.cancel_button4 = CTkButton(self.root, text="Cancel", width=50, height=20, command=lambda: self.cancel_button4_pressed())
        self.cancel_button4.place(x=210, y=350)
        # Create a "Save" button for the days working
        self.save_button4 = CTkButton(self.root, text="Save", width=50, height=20, command=lambda: self.save_button4_pressed())
        self.save_button4.place(x=280, y=350)

    def cancel_button4_pressed(self):
        # Connect to the SQLite database
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        # Retrieve the working days from the database based on the email
        monday_working, tuesday_working, wednesday_working, thursday_working, friday_working, saturday_working, sunday_working = cursor.execute("""SELECT Monday_Working, Tuesday_Working, Wednesday_Working, Thursday_Working,
Friday_Working, Saturday_Working, Sunday_Working FROM Account WHERE Email = (?)""", (self.email,)).fetchall()[0]
        conn.close()
        # Create IntVar instances for the switches based on the retrieved values
        switch_var1 = IntVar(value=monday_working)
        switch_var2 = IntVar(value=tuesday_working)
        switch_var3 = IntVar(value=wednesday_working)
        switch_var4 = IntVar(value=thursday_working)
        switch_var5 = IntVar(value=friday_working)
        switch_var6 = IntVar(value=saturday_working)
        switch_var7 = IntVar(value=sunday_working)
        # Configure the switches with the corresponding IntVar instances and set their state to "disabled"
        self.monday_switch.configure(variable=switch_var1)
        self.monday_switch.configure(state=("disabled"))
        self.tuesday_switch.configure(variable=switch_var2)
        self.tuesday_switch.configure(state=("disabled"))
        self.wednesday_switch.configure(variable=switch_var3)
        self.wednesday_switch.configure(state=("disabled"))
        self.thursday_switch.configure(variable=switch_var4)
        self.thursday_switch.configure(state=("disabled"))
        self.friday_switch.configure(variable=switch_var5)
        self.friday_switch.configure(state=("disabled"))
        self.saturday_switch.configure(variable=switch_var6)
        self.saturday_switch.configure(state=("disabled"))
        self.sunday_switch.configure(variable=switch_var7)
        self.sunday_switch.configure(state=("disabled"))
        # Remove the "Cancel" and "Save" buttons for the days working
        self.cancel_button4.place_forget()
        self.save_button4.place_forget()
        # Re-display the "Edit" button for the days working
        self.edit_button4.place(x=210, y=350)

    def save_button4_pressed(self):
        # Show a confirmation dialog to change the working days
        answer = tk.messagebox.askyesno(title="Confirmation", message="Are you sure you want to change the days you are working?")
        if answer:
            # Get the values of the day switches
            mon = self.monday_switch.get()
            tues = self.tuesday_switch.get()
            wed = self.wednesday_switch.get()
            thur = self.thursday_switch.get()
            fri = self.friday_switch.get()
            sat = self.saturday_switch.get()
            sun = self.sunday_switch.get()
            # Connect to the SQLite database
            conn = sqlite3.connect('Rota_Maker.db')
            cursor = conn.cursor()
            # Update the working days in the database for the current email
            cursor.execute("""UPDATE Account SET Monday_Working = ?, Tuesday_Working = ?, Wednesday_Working = ?, Thursday_Working = ?,
Friday_Working = ?, Saturday_Working = ?, Sunday_Working = ? WHERE Email = ?""", (mon, tues, wed, thur, fri, sat, sun, self.email))
            conn.commit()
            conn.close()
            # Show a success message
            tk.messagebox.showinfo(title="Success", message="Working days have been updated")
            # Call the cancel_button4_pressed method to reset the switches and buttons
            self.cancel_button4_pressed()
        else:
            # If the user cancels, do nothing
            return

    def edit5_pressed(self):
        # Set the state of the contract hours combobox to "readonly"
        self.contract_hrs_combobox.configure(state="readonly")
        # Remove the "Edit" button for the contract hours
        self.edit_button5.place_forget()
        # Create a "Cancel" button for the contract hours
        self.cancel5_button = CTkButton(self.root, text="Cancel", width=50, height=20, command=lambda: self.cancel_button5_pressed())
        self.cancel5_button.place(x=260, y=430)
        # Create a "Save" button for the contract hours
        self.save5_button = CTkButton(self.root, text="Save", width=50, height=20, command=lambda: self.save_button5_pressed())
        self.save5_button.place(x=320, y=430)

    def cancel_button5_pressed(self):
        # Connect to the SQLite database
        conn = sqlite3.connect('Rota_Maker.db')
        cursor = conn.cursor()
        # Retrieve the contract hours from the database based on the email
        contract = cursor.execute("SELECT Contract_Hours FROM Account WHERE Email =(?)", (self.email,)).fetchall()[0][0]
        conn.close()
        # Set the contract hours combobox with the retrieved value
        self.contract_hrs_combobox.set(contract)
        # Set the state of the contract hours combobox to "disabled"
        self.contract_hrs_combobox.configure(state="disabled")
        # Remove the "Cancel" and "Save" buttons for the contract hours
        self.cancel5_button.place_forget()
        self.save5_button.place_forget()
        # Re-display the "Edit" button for the contract hours
        self.edit_button5.place(x=260, y=430)

    def save_button5_pressed(self):
        # Show a confirmation dialog to update the contract hours
        answer = tk.messagebox.askyesno(title="Confirmation", message="Are you sure you want to update your contract hours?")
        if not answer:
            # If the user cancels, do nothing
            return
        else:
            # Get the selected contract hours from the combobox
            hrs = self.contract_hrs_combobox.get()
            # Connect to the SQLite database
            conn = sqlite3.connect("Rota_Maker.db")
            cursor = conn.cursor()
            # Update the contract hours in the database for the current email
            cursor.execute("UPDATE Account SET Contract_Hours =? WHERE Email=?", (hrs, self.email))
            conn.commit()
            conn.close()
            # Show a success message
            tk.messagebox.showinfo(title="Success", message="Contract Hours have been updated")
            # Call the cancel_button5_pressed method to reset the combobox and buttons
            self.cancel_button5_pressed()

    def edit6_pressed(self):
       # Set the state of the primary store combobox to "readonly"
       self.primary_store_combobox.configure(state="readonly")
       # Remove the "Edit" button for the primary store
       self.edit_button6.place_forget()
       # Create a "Cancel" button for the primary store
       self.cancel6_button = CTkButton(self.root, text="Cancel", width=50, height=20, command=lambda: self.cancel_button6_pressed())
       self.cancel6_button.place(x=280, y=490)
       # Create a "Save" button for the primary store
       self.save6_button = CTkButton(self.root, text="Save", width=50, height=20, command=lambda: self.save_button6_pressed())
       self.save6_button.place(x=340, y=490)

    def cancel_button6_pressed(self):
       # Connect to the SQLite database
       conn = sqlite3.connect('Rota_Maker.db')
       cursor = conn.cursor()
       # Retrieve the primary store from the database based on the email
       store = cursor.execute("SELECT Primary_Store FROM Account WHERE Email = (?)", (self.email,)).fetchall()[0][0]
       conn.close()
       # Set the primary store combobox with the retrieved value
       self.primary_store_combobox.set(self.primary_store_val[store-1])
       # Set the state of the primary store combobox to "disabled"
       self.primary_store_combobox.configure(state="disabled")
       # Remove the "Cancel" and "Save" buttons for the primary store
       self.cancel6_button.place_forget()
       self.save6_button.place_forget()
       # Re-display the "Edit" button for the primary store
       self.edit_button6.place(x=280, y=490)

    def save_button6_pressed(self):
        # Show a confirmation dialog to update the current store
       answer = tk.messagebox.askyesno(title="Confirmation", message="Are you sure you want to update your Current Store?")
       if not answer:
           # If the user cancels, do nothing
           return
       else:
           # Get the selected store name from the combobox
           storename = self.primary_store_combobox.get()
           # Create a dictionary to map store names to closest store names
           self.closest_store_dic = {"South Kensington" : "Chiswick", "Battersea" : "East Dulwich", "Hampstead" : "Stoke Newington",
                                     "East Dulwich" : "Battersea", "Chiswick" : "South Kensington", "Stoke Newington" : "Hampstead", "Wimbledon" : "Battersea"}
           # Get the closest store name based on the selected store
           closeststore = self.closest_store_dic[storename]
           # Connect to the SQLite database
           conn = sqlite3.connect('Rota_Maker.db')
           cursor = conn.cursor()
           # Retrieve the store ID and closest store ID based on the store names
           storeid = cursor.execute("SELECT StoreID FROM Stores WHERE Store_Name =?", (storename,)).fetchall()[0][0]
           closeststoreid = cursor.execute("SELECT StoreID FROM Stores WHERE Store_Name =?", (closeststore,)).fetchall()[0][0]
           # Update the primary store and closest store in the database for the current email
           cursor.execute("UPDATE Account SET Primary_Store=?, Closest_Store=? WHERE Email=?", (storeid,closeststoreid, self.email))
           conn.commit()
           conn.close()
           # Show a success message
           tk.messagebox.showinfo(title="Success", message="Current Store has been updated")
           # Call the cancel_button6_pressed method to reset the combobox and buttons
           self.cancel_button6_pressed()

# Check if the script is being run directly (not imported as a module)
if __name__ == "__main__":
    # Create an instance of a Tkinter application window
    root = CTk()
    # Create the first window of the application, passing the root window as an argument
    app = first_window(root)
    # Start the Tkinter event loop
    root.mainloop()
